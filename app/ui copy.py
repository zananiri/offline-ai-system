"""
Gradio UI — talks to the FastAPI backend at localhost:8000.
Run after main.py is already running: python -m app.ui
"""
import base64
import queue
import threading
import unicodedata
import re
import tempfile
import zipfile
from pathlib import Path

import requests
import gradio as gr
import openpyxl
from openpyxl.styles import Font
from pypdf import PdfReader, PdfWriter

try:
    import chromadb
except ImportError:
    chromadb = None  # Canon AI tab degrades to a clear error message if this isn't installed

from app.translate import LANGUAGES
from app.document import chunk_text
# DeepSeek-backed Attorney tabs' grounding pipeline (Attorney 32B / Attorney
# 24B (DeepSeek 14B) -- BGE-M3 + BM25 hybrid retrieval ->
# DeepSeek reasoning -> Dicta draft -> DeepSeek verification loop) -- see
# that module's docstring for the full architecture. Imported under an
# alias so its availability flag reads clearly against LEGAL_MODEL/etc.
# below without colliding with the (unrelated) `chromadb` availability
# flag already used for the Canon/GDPR/HIPAA/Knesset tabs above.
from app.legal_pipeline_v2 import (
    answer_legal_question,
    PIPELINE_AVAILABLE as LEGAL_V2_AVAILABLE,
    DEEPSEEK_MODEL_DEFAULT as LEGAL_V2_DEEPSEEK_32B,
    DEEPSEEK_MODEL_FAST as LEGAL_V2_DEEPSEEK_14B,
)

BACKEND_URL = "http://localhost:8000"

# Backs ONLY the Translate tab's optional "Also summarize with AI" step
# (see process() below). NOTE: this string is duplicated in app/main.py's
# TRANSLATE_SUMMARY_MODEL (same reasoning as LEGAL_MODEL's duplication --
# ui.py only talks to the backend over HTTP and can't share a Python
# constant with it) -- keep the two in sync if you change the model.
# Deliberately kept separate from the backend's general OLLAMA_MODEL
# default (gpt-oss:20b, used by Chat/Canon AI/invoice classification):
# qwen2.5-instruct is explicitly trained for strong multilingual output
# across this app's translation targets, where gpt-oss:20b is more prone
# to drifting back into English despite the "Respond ONLY in {target_lang}"
# instruction below -- see main.py's TRANSLATE_SUMMARY_MODEL comment.
TRANSLATE_SUMMARY_MODEL = "qwen2.5:7b-instruct-q4_K_M"

# Matches (and pads slightly past) main.py's own _OLLAMA_REQUEST_TIMEOUT_SECONDS,
# so this client never gives up before the backend's own safety-net timeout
# would already have returned a clean error. Without this, requests has no
# default timeout at all -- a stuck/slow "thinking" model call just hangs
# the whole Gradio UI indefinitely with zero feedback (confirmed in
# practice: a Clean Air Law question to DictaLM ran 1500+s and 13k+ tokens
# with no sign of stopping before main.py's num_predict cap was added).
_CHAT_TIMEOUT_SECONDS = 1830


def _chat_backend(messages, model=None, num_predict=None, num_ctx=None, timeout=None, timeout_hint=None):
    """
    POSTs to the backend's /chat endpoint and returns the response content
    as a plain string. Raises RuntimeError with a clean, user-facing message
    on any failure (backend unreachable, timed out, or a clean error detail
    the backend itself already generated) -- callers decide what to do with
    that: show it as the reply (chat_fn/legal_chat_fn), or fall back
    gracefully without losing other already-successful work (process()'s
    summarizer step, where a failed summary shouldn't also take down the
    translated text that already succeeded).

    timeout/timeout_hint let a caller override the default for a model that
    needs more room (see legal_chat_fn, which passes a larger timeout and a
    more specific hint tailored to a "thinking" model on a broad question).

    num_ctx lets a caller request a bigger context window than the backend's
    own default -- see main.py's _num_ctx_for comment for why this matters:
    without it, num_predict isn't a real ceiling for a big-num_predict
    caller like legal_chat_fn, since Ollama's smaller default context window
    fills up first and the model keeps going via context-shifting instead of
    stopping cleanly.
    """
    payload = {"messages": messages}
    if model:
        payload["model"] = model
    if num_predict:
        payload["num_predict"] = num_predict
    if num_ctx:
        payload["num_ctx"] = num_ctx
    timeout = timeout or _CHAT_TIMEOUT_SECONDS
    try:
        resp = requests.post(f"{BACKEND_URL}/chat", json=payload, timeout=timeout)
    except requests.exceptions.Timeout:
        hint = timeout_hint or "try a shorter question or a smaller attached document."
        raise RuntimeError(
            f"The model didn't respond within {timeout // 60} minutes. "
            f"It may be stuck, or just very slow on this hardware -- {hint}"
        )
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Couldn't reach the backend: {e}")

    if not resp.ok:
        try:
            detail = resp.json().get("detail", resp.text)
        except Exception:
            detail = resp.text
        raise RuntimeError(detail)

    return resp.json()["content"]


# Model backing the Legal tab. NOTE: this string is duplicated in
# app/main.py (ui.py only talks to that backend over HTTP, so it can't
# import a shared Python constant from it) — keep the two in sync if you
# change the model. Must be pulled once via:
#   ollama pull hf.co/dicta-il/DictaLM-3.0-24B-Thinking-GGUF:Q4_K_M
#
# Upgraded from the 1.7B model to Dicta's 24B flagship for noticeably
# stronger legal reasoning and more reliable citations. Q4_K_M quantization
# is a ~14.3GB download/on-disk file. On a 32GB-RAM, offline/CPU (or
# modest-iGPU) machine this fits with room to spare for the OS and the rest
# of this app's own memory use (docling, the MADLAD-400 translation model,
# etc.) -- but it's the biggest single thing this app loads, so:
#   - Real tradeoff vs. the 1.7B: much slower generation (CPU tok/s roughly
#     tracks parameter count, so expect noticeably fewer tokens/sec than the
#     1.7B's ~20 tok/s -- see the timeout constants below, which were raised
#     accordingly).
#   - If you're running the Legal tab at the same time as a large translation
#     or batch-invoice job (i.e. multiple big models resident at once) and
#     see swapping/OOM, drop to the smaller IQ4_XS quant instead (~12.8GB:
#     hf.co/dicta-il/DictaLM-3.0-24B-Thinking-GGUF:IQ4_XS), or fall back to
#     the 1.7B line below.
# Previous (faster, weaker) setting, kept here for an easy revert:
#   LEGAL_MODEL = "hf.co/dicta-il/DictaLM-3.0-1.7B-Thinking-GGUF:Q4_K_M"
LEGAL_MODEL = "hf.co/dicta-il/DictaLM-3.0-24B-Thinking-GGUF:Q4_K_M"

# See main.py's _DEFAULT_NUM_PREDICT comment for the full story: with no cap
# at all, a "thinking" model on a broad question can run for as long as it
# wants (DictaLM-3.0-24B-Thinking has a native 65k-token context, so there's
# no natural ceiling either -- confirmed in practice with the 1.7B model: a
# real question about the Clean Air Law ran 1500+s and 13k+ tokens with no
# end in sight, and the 24B is slower still). This is more generous than the
# backend's own 4096-token default since thinking + a citation for every
# claim genuinely needs more room, but it's still a hard cap -- if it's hit
# mid-thought, _strip_thinking (main.py) turns that into a clear "ran out of
# space" message instead of hanging forever.
_LEGAL_NUM_PREDICT = 6144
# _LEGAL_NUM_CTX only works as an actual ceiling if the model's context
# window is at least that big -- see main.py's _num_ctx_for comment. Sized
# for num_predict (6144) + the system prompt + chat history + a full
# MAX_CONTEXT_CHARS-sized attached document, with headroom -- comfortably
# under DictaLM-3.0-24B-Thinking's native 65k.
#
# CORRECTED to 11264 (was left at 16384 in this file despite main.py's own
# comment already saying it should be 11264 -- the two had drifted out of
# sync, which is the actual root cause of "the Attorney tab answer never
# comes back, no error" on a 32GB-RAM/no-GPU-offload machine: RAM is
# already ~85% used at idle on this hardware, and a 24B model's KV cache at
# 16384 tokens pushes total memory use past what's physically free,
# triggering pagefile-swapped CPU inference -- which doesn't error, it just
# runs enormously slower than normal while looking identical to a genuine
# hang. This is exactly why _KNESSET_WHOLE_LAW_MAX_WEIGHT below is sized
# against THIS number, not the stale 16384.
#
# (Note: KV cache at this context size still adds real memory on top of
# the 14.3GB weights -- a few more GB. Lower further if you're tight on RAM
# alongside other models this app loads. Also make sure you're starting
# Ollama via run.ps1/setup.ps1/gui_run.py, not a bare `ollama serve` -- those
# scripts set OLLAMA_KV_CACHE_TYPE=q8_0, which roughly halves KV-cache
# memory and is part of what makes this context size affordable at all on
# this hardware.)
_LEGAL_NUM_CTX = 11264
# Comfortably above the backend's own safety-net timeout (main.py's
# _OLLAMA_REQUEST_TIMEOUT_SECONDS), so that backend's clearer error message
# surfaces first instead of a generic timeout here. Raised well past the
# 1.7B-era value: on CPU, a 24B dense model generating up to 6144 tokens can
# genuinely take tens of minutes rather than single-digit minutes.
_LEGAL_REQUEST_TIMEOUT_SECONDS = 3700

# --- "Attorney 1.7B (Fast)" tab: same LEGAL_SYSTEM_PROMPT and Knesset RAG
# grounding, swapped to Dicta's much smaller 1.7B model for CPU-bound
# hardware where the 24B's per-question latency (documented above as
# "tens of minutes... genuinely take tens of minutes") isn't workable. This
# is a real speed/RAM win, not a free one -- see the tab's own gr.Markdown
# description below and _LEGAL_FAST_MODEL's warning docstring for the
# quality tradeoff: DictaLM-3.0-1.7B-Thinking is a materially weaker legal
# reasoner than the 24B and more prone to fabricating specific citations,
# which is exactly why the hallucination safety nets added above
# (_KNESSET_DISCLAIMER_UNGROUNDED_WITH_CITATIONS etc.) matter MORE here,
# not less -- they're reused as-is for this tab, unchanged.
#
# Must be pulled once via:
#   ollama pull hf.co/dicta-il/DictaLM-3.0-1.7B-Thinking-GGUF:Q4_K_M
LEGAL_MODEL_FAST = "hf.co/dicta-il/DictaLM-3.0-1.7B-Thinking-GGUF:Q4_K_M"

# A 1.7B model's weights (~1-2GB at Q4_K_M) and per-token KV-cache footprint
# are both a small fraction of the 24B's -- the RAM-pressure/pagefile-swap
# problem that forced _LEGAL_NUM_CTX down to 11264 above essentially
# shouldn't apply here, so this can afford a much larger context window on
# the SAME 32GB/no-GPU hardware. That in turn means whole-law Knesset
# retrieval (see _KNESSET_FAST_WHOLE_LAW_MAX_WEIGHT below) can include a
# meaningfully bigger excerpt before truncating. These are reasonable
# starting numbers, not independently benchmarked on this exact model --
# watch the same [Knesset RAG] log line and `ollama ps` memory/CPU% this
# app already uses for the 24B tab, and adjust if your hardware disagrees.
_LEGAL_FAST_NUM_PREDICT = 4096
_LEGAL_FAST_NUM_CTX = 24576
# Raised 900 -> 1800 (30 min): 15 min turned out too tight in practice --
# a real request hit this timeout with no evidence in the FastAPI/Gradio
# console of anything actually stuck (retrieval/rerank completed normally
# right before it), meaning the 1.7B model itself was just still generating
# when the clock ran out, not hung. Still kept well under the 24B tab's
# 3700s and the backend's own 3600s safety net (main.py's
# _OLLAMA_REQUEST_TIMEOUT_SECONDS), so a genuinely stuck request still
# fails well before an hour, just with more realistic headroom for normal
# slow-hardware generation than 900s gave it.
_LEGAL_FAST_REQUEST_TIMEOUT_SECONDS = 1800

# Whole-law budget for the fast tab -- see _KNESSET_WHOLE_LAW_MAX_WEIGHT's
# comment for the full reasoning behind how this number is derived
# (available input budget = num_ctx - num_predict - ~1500 tokens headroom).
# With _LEGAL_FAST_NUM_CTX=24576 and _LEGAL_FAST_NUM_PREDICT=4096, that's
# roughly 19000 tokens of headroom -- generously sized at 8000 weighted
# units so most laws fit whole without truncating, while still leaving a
# real margin rather than using every last token of the window.
_KNESSET_FAST_WHOLE_LAW_MAX_WEIGHT = 8000

LEGAL_SYSTEM_PROMPT = (
    "You are an Israeli lawyer. Think through and answer every question strictly "
    "according to the laws of the State of Israel -- its statutes, regulations, "
    "and case law -- not the law of any other jurisdiction, unless the user "
    "explicitly asks about a different country's law.\n\n"
    "For every substantive legal claim, name the specific Israeli statute, "
    "regulation, or section you are relying on immediately after the claim -- "
    "for example: 'A contract requires offer and acceptance (Section 1, "
    "Contracts Law (General Part), 5733-1973).' Also cite the relevant "
    "section/clause of any attached document when you rely on it. If you are "
    "not confident of the exact statute, section number, or case citation, say "
    "so explicitly instead of inventing one -- a wrong or fabricated citation "
    "is worse than admitting uncertainty.\n\n"
    "You may also be given retrieved statute excerpts below, pulled "
    "directly from the Knesset's official legislation database. When they "
    "are relevant to the question, prefer citing and quoting from those "
    "exact excerpts (including the סעיף/section number given) over relying "
    "on general recall -- they are authoritative and more current than "
    "your training data. If no excerpts were retrieved, or the retrieved "
    "excerpts don't actually cover the question, say so explicitly rather "
    "than presenting an uncited claim as if it came from them.\n\n"
    "Keep your answer proportionate to the question: for a broad or general "
    "topic, cover the most important, directly relevant points rather than "
    "exhaustively enumerating every provision of a law -- you can offer to go "
    "deeper on a specific part if the user wants that.\n\n"
    "You are not a substitute for advice from a licensed attorney, and you "
    "should say so when a question calls for one. Always reply in the same "
    "language the user's question is written in -- e.g. if they write in "
    "English, your entire answer must be in English, even though the "
    "underlying law and any attached document may be in Hebrew. Never switch "
    "languages on the user unasked."
)

# Best-effort heuristic for "did the answer cite anything at all" -- matches
# common Israeli-law citation shapes in both English and Hebrew (section/
# regulation numbers, named statutes, and Hebrew court-ruling abbreviations
# like בג"ץ/ע"א). This can only detect the ABSENCE of a citation-shaped
# string; it cannot verify that a citation that IS present is real. Small
# models are prone to fabricating plausible-looking statute/section numbers,
# and no regex can catch that -- this is a floor (something was cited),
# not a correctness guarantee. See legal_chat_fn for how it's used.
_LEGAL_CITATION_RE = re.compile(
    r"(סעיף\s*\d+|תקנה\s*\d+|חוק\s+\S+|פסק\s*דין|בג\"?ץ|ע\"?א\s*\d+|"
    r"\bsection\s+\d+|\bregulation\s+\d+|\barticle\s+\d+|\blaw\s*(\(|,)?\s*(19|20|5[6-9])\d{2}\b|"
    r"\bhcj\b|\bbasic\s+law\b)",
    re.IGNORECASE,
)

_NO_CITATION_NOTE = (
    "\n\n---\n⚠️ *This answer doesn't appear to cite a specific law, regulation, "
    "or case. Treat it as general information only and verify against the "
    "actual legislation or with a licensed attorney before relying on it.*"
)

# Splits an answer into sentence-ish chunks so _extract_citations can pull
# out the whole sentence around a citation match (not just the bare
# "Section 1" fragment _LEGAL_CITATION_RE matches on its own) -- that's what
# actually gets shown in the sidebar panel. Handles Hebrew sentence-enders
# too, plus plain newlines (the model often puts one citation per line).
_SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+|\n+")

_CITATIONS_PLACEHOLDER = (
    "_No citations detected yet. Once you ask a question, any statute, "
    "regulation, or case citation the model includes in its answer will be "
    "listed here for easy review._"
)
_CITATIONS_NONE_FOUND = (
    "_This answer didn't include anything citation-shaped. See the "
    "warning in the chat -- treat it as general information only._"
)


def _extract_citations(answer: str) -> list[str]:
    """Best-effort pull of citation-bearing sentences out of an answer, for
    display in the Legal tab's sidebar. Same caveats as _LEGAL_CITATION_RE:
    this can only detect something citation-shaped, not verify it's real."""
    seen = set()
    citations = []
    for chunk in _SENTENCE_SPLIT_RE.split(answer):
        sentence = chunk.strip(" -•\t\n")
        if not sentence or sentence in seen:
            continue
        if _LEGAL_CITATION_RE.search(sentence):
            seen.add(sentence)
            citations.append(sentence)
    return citations


def _format_citations_panel(citations: list[str]) -> str:
    if not citations:
        return _CITATIONS_NONE_FOUND
    return "\n\n".join(f"- {c}" for c in citations)

# Languages whose script reads right-to-left. Used to flip the translated-text
# output box's text direction so Arabic/Hebrew results display correctly
# instead of being left-aligned like Latin-script languages.
RTL_LANGUAGES = {"arabic", "hebrew"}


def _is_rtl(lang: str) -> bool:
    return (lang or "").strip().lower() in RTL_LANGUAGES


_RTL_MARK = "\u200f"  # RIGHT-TO-LEFT MARK (invisible, sets bidi direction only)


def _anchor_rtl_lines(text: str) -> str:
    """
    Textbox(rtl=True) sets the box's overall base direction, but that alone
    doesn't stop individual lines from getting visually reordered by the
    browser's bidi algorithm -- a line that happens to start with a digit,
    punctuation, or an embedded Latin word/number (invoice numbers, dates,
    stray English terms) can pull that whole line's layout toward
    left-to-right, scattering words out of their intended order.

    Prefixing every non-empty line with an invisible RIGHT-TO-LEFT MARK
    (U+200F) fixes each line's base direction as RTL regardless of its
    first character, without adding any visible content.
    """
    if not text:
        return text
    return "\n".join(
        f"{_RTL_MARK}{line}" if line.strip() else line
        for line in text.split("\n")
    )


# --- Document preview (Translate tab's right-hand panel) -------------------
#
# There's no single Gradio component that previews every format this app
# accepts, so three different strategies are used depending on file type:
#
#   Images  -> gr.Image with show_fullscreen_button=True. Gives a native
#              click-to-zoom/pan lightbox for free (same mechanism this
#              file already uses for the header logo, just switched on).
#   PDFs    -> embedded via a base64 data: URI inside an <iframe>, so the
#              browser's own PDF viewer renders it -- built-in zoom, scroll,
#              page navigation, and print, no extra dependency. A data URI
#              is used instead of a path-based Gradio static-file URL (e.g.
#              "/file=...") on purpose: that route's exact prefix differs
#              across Gradio major versions and only works for paths inside
#              Gradio's allowed-paths, which is fragile to hardcode. A data
#              URI works identically regardless of Gradio version.
#   Other   -> (DOCX, PPTX, etc.) no in-browser renderer exists for these
#              without extra dependencies, so the existing /extract-text
#              endpoint is reused to show the extracted text instead. Not a
#              visual preview, but it directly shows what's about to be
#              translated, and costs nothing extra: DOCX/PPTX never go
#              through OCR (see document.py), so this second extraction
#              call is fast.
_PREVIEW_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}

# Above this, base64-inlining the whole file into the page's HTML would
# bloat the page enough to feel sluggish -- fall back to a plain notice
# instead of a broken/slow preview. Translation itself is unaffected either
# way; this only gates the preview panel.
_MAX_INLINE_PREVIEW_BYTES = 20 * 1024 * 1024  # 20 MB


def _pdf_preview_html(path: Path) -> str:
    b64 = base64.b64encode(path.read_bytes()).decode("ascii")
    data_uri = f"data:application/pdf;base64,{b64}"
    return f"""
<div style="border:1px solid var(--border-color-primary, #ddd); border-radius:8px; overflow:hidden;">
  <iframe src="{data_uri}" style="width:100%; height:520px; border:none;"></iframe>
</div>
<p style="text-align:center; margin-top:6px;">
  <a href="{data_uri}" target="_blank" rel="noopener">🔍 Open full preview in a new tab (zoom, scroll, print)</a>
</p>
"""


def build_preview(file):
    """
    Builds a live preview of the uploaded document for the panel on the
    right of the Translate tab. Wired to file_in.change, so it fires the
    moment a file is picked (or cleared) -- independent of clicking
    Translate -- letting the user confirm it's the right document before
    running anything.

    The whole preview column is hidden by default and only made visible
    once a file is actually present, and hidden again when the file is
    cleared -- there's no "upload a document" placeholder state, the panel
    simply isn't there until there's something to show.

    Returns a 4-tuple of gr.update(...) for (preview_column, image_preview,
    pdf_preview_html, text_preview). Within the column, exactly one of the
    three preview components is made visible at a time (text_preview also
    doubles as the spot for status/error messages, e.g. an oversized PDF
    or a failed extraction).
    """
    hidden = gr.update(visible=False)

    if file is None:
        return gr.update(visible=False), hidden, hidden, hidden

    path = Path(file.name)
    suffix = path.suffix.lower()

    if suffix in _PREVIEW_IMAGE_EXTS:
        return gr.update(visible=True), gr.update(value=str(path), visible=True), hidden, hidden

    if suffix == ".pdf":
        size = path.stat().st_size
        if size > _MAX_INLINE_PREVIEW_BYTES:
            msg = (
                f"⚠️ This PDF is {size / 1_048_576:.1f} MB, too large to preview inline. "
                "It will still be translated normally -- the preview panel just skips it."
            )
            return gr.update(visible=True), hidden, hidden, gr.update(value=msg, visible=True)
        return gr.update(visible=True), hidden, gr.update(value=_pdf_preview_html(path), visible=True), hidden

    # DOCX/PPTX/etc. -- no visual renderer available, fall back to extracted text.
    try:
        with open(path, "rb") as f:
            resp = requests.post(f"{BACKEND_URL}/extract-text", files={"file": f}, data={"hebrew": False})
        resp.raise_for_status()
        text = resp.json().get("markdown", "")
    except Exception as e:
        return gr.update(visible=True), hidden, hidden, gr.update(value=f"⚠️ Couldn't build a preview: {e}", visible=True)

    if not text.strip():
        return gr.update(visible=True), hidden, hidden, gr.update(value="(no previewable text found in this document)", visible=True)
    return gr.update(visible=True), hidden, hidden, gr.update(value=text, visible=True)


def process(file, source_lang, target_lang, summarize, hebrew_doc, progress=gr.Progress()):
    progress(0, desc="Extracting text from document (OCR if needed)...")
    with open(file.name, "rb") as f:
        extract_resp = requests.post(
            f"{BACKEND_URL}/extract-text", files={"file": f}, data={"hebrew": hebrew_doc}
        )
    extract_resp.raise_for_status()
    extract_data = extract_resp.json()
    markdown_text = extract_data["markdown"]
    # The backend can auto-detect Hebrew and route through Tesseract even
    # if hebrew_doc was False (see document.py's resolve_hebrew_flag) --
    # report what actually ran, not just what was requested, so this
    # doesn't silently lie when that override kicks in.
    hebrew_used = extract_data.get("hebrew_used", hebrew_doc)
    ocr_engine = "Tesseract (Hebrew)" if hebrew_used else "RapidOCR (default)"

    chunks = chunk_text(markdown_text)
    if not chunks:
        return (
            gr.update(value="(no text found in document)", rtl=_is_rtl(target_lang)),
            ocr_engine,
            "(not requested)",
        )

    total_steps = len(chunks) + (1 if summarize else 0)
    translated_chunks = []
    failed_chunks = 0
    for i, chunk in enumerate(chunks):
        progress((i + 1) / total_steps, desc=f"Translating chunk {i + 1}/{len(chunks)}...")
        resp = requests.post(
            f"{BACKEND_URL}/translate-chunk",
            json={"text": chunk, "source_lang": source_lang, "target_lang": target_lang},
        )
        resp.raise_for_status()
        data = resp.json()
        translated_value = data.get("translated", "")
        ok = data.get("ok", True)
        if not isinstance(translated_value, str):
            # Defensive: a backend/frontend version mismatch (e.g. main.py
            # not unpacking the (text, ok) tuple translate() now returns)
            # would otherwise land a list/tuple here and crash the whole
            # request at the join() below. Degrade to a visible failure
            # instead of a hard crash.
            translated_value = str(translated_value)
            ok = False
        if ok:
            translated_chunks.append(translated_value)
        else:
            failed_chunks += 1
            translated_chunks.append(
                f">>> COULD NOT TRANSLATE THIS SECTION (shown untranslated below) >>>\n"
                f"{translated_value}\n"
                f"<<< END UNTRANSLATED SECTION <<<"
            )

    translated_text = "\n\n".join(translated_chunks)
    if failed_chunks:
        translated_text = (
            f"⚠️ {failed_chunks} of {len(chunks)} section(s) could not be translated and are shown "
            f"in their ORIGINAL, untranslated form below (marked with >>> / <<<). This almost always "
            f"means the extracted text for those sections was garbled — usually poor OCR quality on "
            f"a scanned page, rather than a translation problem. Check the OCR engine used below, and "
            f"whether the source document/scan quality is high enough.\n\n"
        ) + translated_text

    summary = "(not requested)"
    if summarize:
        progress(1.0, desc="Summarizing with AI...")
        messages = [
            {
                "role": "system",
                "content": (
                    f"You are a precise document summarizer. Respond ONLY in "
                    f"{target_lang}, matching the language of the text you are "
                    f"given -- never switch to a different language."
                ),
            },
            {"role": "user", "content": f"Summarize this in 3-5 bullet points:\n\n{translated_text}"},
        ]
        try:
            summary = _chat_backend(messages, model=TRANSLATE_SUMMARY_MODEL)
        except RuntimeError as e:
            summary = f"⚠️ Summary failed: {e}"

    is_rtl = _is_rtl(target_lang)
    display_text = _anchor_rtl_lines(translated_text) if is_rtl else translated_text
    display_summary = (
        gr.update(value=_anchor_rtl_lines(summary), rtl=True) if (is_rtl and summarize)
        else gr.update(value=summary, rtl=False)
    )
    return gr.update(value=display_text, rtl=is_rtl), ocr_engine, display_summary


def convert_to_word(file, hebrew_doc, progress=gr.Progress()):
    progress(0.15, desc="Extracting text (running OCR if needed)...")
    with open(file.name, "rb") as f:
        resp = requests.post(
            f"{BACKEND_URL}/convert-to-word", files={"file": f}, data={"hebrew": hebrew_doc}
        )
    resp.raise_for_status()
    progress(0.9, desc="Saving Word document...")

    out_name = Path(file.name).stem + ".docx"
    out_path = str(Path(tempfile.gettempdir()) / out_name)
    with open(out_path, "wb") as out_f:
        out_f.write(resp.content)
    progress(1.0, desc="Done")

    # Auto-detection (document.py's resolve_hebrew_flag) can route this
    # through Tesseract even if hebrew_doc was left unchecked -- surfaced
    # via a response header since FileResponse can't carry a JSON body.
    hebrew_used = resp.headers.get("X-Hebrew-OCR-Used", str(hebrew_doc)) == "True"
    ocr_engine = "Tesseract (Hebrew)" if hebrew_used else "RapidOCR (default)"
    return out_path, ocr_engine


MAX_CONTEXT_CHARS = 6000  # keep injected document text within the model's comfortable context window


def extract_context_from_files(filepaths, hebrew=False):
    contexts = []
    for path in filepaths:
        with open(path, "rb") as f:
            resp = requests.post(
                f"{BACKEND_URL}/extract-text", files={"file": f}, data={"hebrew": hebrew}
            )
        resp.raise_for_status()
        data = resp.json()
        contexts.append(f"--- Content of {Path(path).name} ---\n{data['markdown']}")
    return "\n\n".join(contexts)


# Matches a request to CREATE a presentation, e.g. "make me a powerpoint
# about X", "generate a slide deck on Y", "can you build a pptx for Z".
# Requires both a presentation-ish noun AND a creation verb, so it doesn't
# fire on unrelated mentions of the word "presentation" or "slides" (e.g.
# "what should I say in my presentation tomorrow?").
_PPTX_NOUN_RE = re.compile(r"\b(power ?point|pptx|slide ?deck|slides?|presentation)\b", re.IGNORECASE)
_PPTX_VERB_RE = re.compile(
    r"\b(make|create|generate|build|write|prepare|put together|draft|produce)\b", re.IGNORECASE
)


def _is_pptx_request(text: str) -> bool:
    text = text or ""
    return bool(_PPTX_NOUN_RE.search(text)) and bool(_PPTX_VERB_RE.search(text))


# Matches a request to rewrite/polish/formalize text, e.g. "rewrite this
# more professionally", "polish the language", "make this sound formal".
# When this fires on a message with an attached file, chat_fn routes to
# rewrite_document_professionally() instead of the normal single-shot chat
# call -- see that function's docstring for why a single call can't do this
# faithfully for a long document.
_REWRITE_TRIGGER_RE = re.compile(
    r"\b(re-?write|polish|proofread|professionali[sz]e|tidy up|clean up|"
    r"formali[sz]e)\b|\bmore professional\b|\bmake.{0,25}professional\b",
    re.IGNORECASE,
)


def _is_rewrite_request(text: str) -> bool:
    return bool(_REWRITE_TRIGGER_RE.search(text or ""))


# Paragraph-sized, not sentence-sized like translate.py's chunk_text(max_chars=400)
# -- a rewrite call needs enough surrounding context per chunk to produce
# coherent, professional prose, but still small enough that num_predict
# below comfortably covers the rewritten output without the model needing
# to compress/summarize to fit.
_REWRITE_CHUNK_CHARS = 3000

# Rewritten prose is rarely much longer than the original for a given
# chunk size, so this headroom (well above _REWRITE_CHUNK_CHARS in tokens)
# is enough to let each chunk come back in full rather than getting cut off
# or compressed. Kept well under _DEFAULT_NUM_CTX (8192) so the backend's
# default num_ctx doesn't need overriding per call.
_REWRITE_NUM_PREDICT_PER_CHUNK = 1600

_REWRITE_SYSTEM_PROMPT = (
    "You are a professional editor. Rewrite the user's text in clear, "
    "professional English. Preserve every fact, figure, name, date, and "
    "detail exactly as given -- do NOT summarize, shorten, condense into a "
    "table or bullet points, or omit anything. Keep the same structure, "
    "order, and level of detail as the original; only improve grammar, "
    "tone, and word choice. Output ONLY the rewritten text, with no "
    "preamble, headers, or commentary of your own."
)


def rewrite_document_professionally(file_context: str) -> str:
    """
    Rewrites a (potentially long) attached document into professional
    English while preserving all information, by processing it in
    paragraph-sized chunks and reassembling the results -- mirroring the
    per-chunk approach translate.py already uses for the same reason: a
    single call on a whole multi-page document runs into the model's
    num_predict/num_ctx ceiling and starts compressing/summarizing instead
    of reproducing everything, which is exactly the "13 pages -> 1-2 page
    table" failure this was written to fix.
    """
    chunks = chunk_text(file_context, max_chars=_REWRITE_CHUNK_CHARS)
    rewritten_parts = []
    for chunk in chunks:
        messages = [
            {"role": "system", "content": _REWRITE_SYSTEM_PROMPT},
            {"role": "user", "content": chunk},
        ]
        rewritten_parts.append(
            _chat_backend(messages, num_predict=_REWRITE_NUM_PREDICT_PER_CHUNK)
        )
    return "\n\n".join(rewritten_parts)


def generate_presentation(prompt: str) -> str:
    """Calls the backend's /generate-pptx endpoint and saves the returned
    file to a fresh temp directory (per-call, so concurrent chats can't
    clobber each other's presentation.pptx)."""
    resp = requests.post(f"{BACKEND_URL}/generate-pptx", json={"prompt": prompt}, timeout=300)
    resp.raise_for_status()
    out_path = str(Path(tempfile.mkdtemp()) / "presentation.pptx")
    with open(out_path, "wb") as f:
        f.write(resp.content)
    return out_path


def chat_fn(message, history, hebrew_doc=False):
    """
    message is a dict {"text": str, "files": [filepaths]} because the
    ChatInterface below is configured with multimodal=True. hebrew_doc
    comes from the additional_inputs checkbox added to the ChatInterface.
    """
    if isinstance(message, dict):
        user_text = message.get("text", "")
        files = message.get("files", []) or []
    else:
        user_text = message
        files = []

    is_rewrite = _is_rewrite_request(user_text)

    file_context = ""
    if files:
        yield "📄 *Extracting text from the attached file(s)…*"
        file_context = extract_context_from_files(files, hebrew=hebrew_doc)
        # A rewrite request needs the WHOLE document, not a 6000-char
        # preview -- truncating here was the main reason a 13-page document
        # only ever came back as 1-2 pages: everything past ~6000 chars
        # never even reached the model. Only truncate for normal Q&A-style
        # attachments, where a preview is a reasonable tradeoff.
        if not is_rewrite and len(file_context) > MAX_CONTEXT_CHARS:
            file_context = file_context[:MAX_CONTEXT_CHARS] + "\n[...truncated, file is longer...]"

    if is_rewrite and file_context:
        yield "✍️ *Rewriting the document in professional English (this can take a while for long documents)…*"
        try:
            yield rewrite_document_professionally(file_context)
        except RuntimeError as e:
            yield f"⚠️ {e}"
        return

    # PowerPoint generation is handled as its own branch rather than folded
    # into the normal /chat call: it needs a structured JSON outline from
    # the model (see app/pptx_generator.py), then a real .pptx file built
    # from that outline and returned as a download, not a chat reply string.
    if _is_pptx_request(user_text):
        prompt = user_text
        if file_context:
            prompt = f"{user_text}\n\nBase the slides on this source material:\n{file_context}"
        yield "🖌️ *Building your PowerPoint outline and slides…*"
        try:
            pptx_path = generate_presentation(prompt)
        except requests.HTTPError as e:
            yield (
                "Sorry, I couldn't generate that presentation "
                f"({e}). Try rephrasing the topic, or try again."
            )
            return
        yield [
            "Here's the presentation you asked for — click below to download it:",
            gr.File(pptx_path),
        ]
        return

    # Keep only plain-text turns from history — earlier attached files aren't
    # re-sent each turn (they already informed the answer they were attached to).
    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    if file_context:
        combined_message = (
            f"The user attached the following document(s):\n\n{file_context}\n\n"
            f"User question: {user_text}"
        )
    else:
        combined_message = user_text

    messages = clean_history + [{"role": "user", "content": combined_message}]
    yield "🤖 *Thinking…*"
    try:
        yield _chat_backend(messages)
    except RuntimeError as e:
        yield f"⚠️ {e}"


def _annotate_legal_answer(answer: str, n_found: int) -> tuple[str, list[str]]:
    """
    Appends the citation-detection warning and Knesset-grounding disclaimer
    to a raw model answer. Shared by both the chunked whole-law path
    (_answer_whole_law_chunked, via _legal_chat_fn_impl) and the normal
    single-call path further down in _legal_chat_fn_impl, so the same
    citation/hallucination safety nets apply no matter which retrieval path
    produced the answer -- factored out specifically so the two paths can't
    silently drift out of sync with each other.

    Returns (annotated_answer, citations) -- citations is exposed so the
    caller can render the "Citations found" sidebar panel without needing
    to re-run _extract_citations a second time on the now-annotated text.
    """
    citations = _extract_citations(answer)
    if not citations:
        answer += _NO_CITATION_NOTE

    # The dangerous combination isn't "no citation" or "no retrieval" alone
    # -- it's BOTH signals at once: nothing was retrieved to ground this
    # answer, yet the model still produced text that LOOKS like a specific
    # legal citation. _extract_citations can't tell a real citation from a
    # fabricated one (see its own docstring), and n_found==0 can't tell you
    # whether the model then invented something anyway -- only checking
    # BOTH together catches this specific, highest-risk case, so it gets a
    # sharper, distinct warning instead of the generic "ungrounded" note.
    if n_found == 0 and citations:
        answer += _KNESSET_DISCLAIMER_UNGROUNDED_WITH_CITATIONS
    else:
        answer += _KNESSET_DISCLAIMER_GROUNDED if n_found else _KNESSET_DISCLAIMER_UNGROUNDED

    return answer, citations


def _legal_chat_fn_impl(
    message, history, hebrew_doc,
    *, model: str, num_predict: int, num_ctx: int, request_timeout: int,
    whole_law_max_weight: float,
):
    """
    Shared implementation behind both Attorney tabs (24B and 1.7B) --
    RAG-grounded on Israeli statute text retrieved from the Knesset
    OData-sourced ChromaDB collection (see retrieve_knesset_laws above --
    either the whole text of a specifically-named law, or a top-K semantic
    search, depending on the question; see that function's docstring).

    Every tab-specific model/context/timeout/budget setting is passed in
    explicitly rather than read from a single shared constant, so the two
    tabs can genuinely run different-sized models with correctly-sized
    context windows at the same time -- see legal_chat_fn / legal_chat_fn_fast
    below for the concrete numbers each tab uses.

    Returns (answer, citations_panel_markdown, knesset_sources_panel_markdown)
    -- three outputs via ChatInterface's additional_outputs: the existing
    regex-based "did the answer cite anything" panel, plus a panel showing
    what was actually retrieved from the law database for this question
    (distinct things: the citations panel can't tell a real citation from
    a fabricated one; the sources panel shows what was actually fed to the
    model as grounding).

    Every returned answer -- grounded or not -- gets a disclaimer footer
    appended (_KNESSET_DISCLAIMER_GROUNDED / _KNESSET_DISCLAIMER_UNGROUNDED /
    _KNESSET_DISCLAIMER_UNGROUNDED_WITH_CITATIONS), on top of the existing
    _NO_CITATION_NOTE. These checks are model-agnostic and unchanged
    regardless of which tab calls this -- if anything they matter MORE on
    the smaller/faster model, which is more prone to fabricating citations,
    not less.
    """
    if isinstance(message, dict):
        user_text = message.get("text", "")
        files = message.get("files", []) or []
    else:
        user_text = message
        files = []

    file_context = ""
    if files:
        yield "📄 *Extracting text from the attached file(s)…*", gr.skip(), gr.skip()
        file_context = extract_context_from_files(files, hebrew=hebrew_doc)
        if len(file_context) > MAX_CONTEXT_CHARS:
            file_context = file_context[:MAX_CONTEXT_CHARS] + "\n[...truncated, file is longer...]"

    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    # --- Whole-law path: if the question clearly names one specific law
    # (see _detect_named_law), read the ENTIRE law via chunked map-reduce
    # (_answer_whole_law_chunked) instead of stuffing it into one oversized
    # prompt -- see that function's docstring / _MAP_BATCH_MAX_WEIGHT's
    # comment for why. Falls through to ordinary (now relevance-floored)
    # semantic search below if no law is confidently named, the fetch
    # fails, or every batch comes back irrelevant to the actual question.
    yield "📚 *Searching the Knesset legislation database…*", gr.skip(), gr.skip()
    knesset_collection = _get_knesset_collection()
    named = (
        _detect_named_law(user_text, _get_knesset_law_titles(knesset_collection))
        if knesset_collection is not None else None
    )

    if named:
        law_id, title, score = named
        print(f"[Knesset RAG] Question names a specific law ({title!r}, token-overlap "
              f"score={score:.2f}) -- using chunked whole-law map-reduce instead of top-K search.")
        sections = _fetch_whole_law_sections(knesset_collection, law_id, title)
        if sections:
            yield (f"⚖️ *Found {title} ({len(sections)} section(s)) — reading it in batches…*",
                   gr.skip(), gr.skip())
            chunked_answer, chunked_sources, chunked_info = yield from _answer_whole_law_chunked(
                user_text, title, sections, clean_history, file_context,
                model=model, num_predict=num_predict, num_ctx=num_ctx,
                request_timeout=request_timeout, reduce_max_weight=whole_law_max_weight,
            )
            if chunked_answer:
                n_found = len(chunked_sources) if chunked_sources else 0
                chunked_answer, chunked_citations = _annotate_legal_answer(chunked_answer, n_found)
                yield (
                    chunked_answer,
                    _format_citations_panel(chunked_citations),
                    _format_knesset_sources_panel(chunked_sources, chunked_info),
                )
                return
            print(f"[Knesset RAG] {title!r}: chunked map-reduce found nothing relevant to "
                  "this question -- falling back to semantic search.")

    law_context, law_sources, retrieval_info = retrieve_knesset_laws(
        user_text, whole_law_max_weight=whole_law_max_weight
    )
    n_found = len(law_sources) if law_sources else 0
    if law_context is None:
        # DB unavailable/unreachable -- degrade gracefully, don't block the
        # tab. Distinguish WHICH service failed (see retrieve_knesset_laws'
        # docstring) so this is diagnosable from the chat itself, not just
        # by reading server-side logs.
        reason = (retrieval_info or {}).get("reason")
        if reason == "embedding_unavailable":
            status = ("⚖️ *Couldn't reach the embedding model (nomic-embed-text via Ollama) "
                       "to search the Knesset database — answering from the Agent's own knowledge…*")
        elif reason == "chroma_unavailable":
            status = ("⚖️ *Couldn't open the Knesset vector database (ChromaDB) — "
                       "answering from the Agent's own knowledge…*")
        else:
            status = "⚖️ *Knesset database unavailable — answering from the Agent's own knowledge…*"
    elif n_found:
        status = f"⚖️ *Found {n_found} relevant statute excerpt(s) — consulting AI…*"
    else:
        status = "⚖️ *Nothing matched in the Knesset database — consulting AI…*"
    yield status, gr.skip(), gr.skip()

    message_parts = []
    if law_context:
        message_parts.append(
            f"Retrieved Israeli statute excerpts relevant to the question "
            f"(from the Knesset's official legislation database):\n\n{law_context}"
        )
    else:
        # Distinct from LEGAL_SYSTEM_PROMPT's general "say so instead of
        # inventing" instruction: that line sits once at the top of a long
        # system prompt and competes with everything else in it for the
        # model's attention. Repeating a sharper, question-specific version
        # of the same instruction right next to THIS question -- when we
        # positively know retrieval found nothing -- is more likely to
        # actually be followed by a locally-quantized model than relying on
        # the system prompt alone. This is a mitigation, not a guarantee;
        # see the combined-signal check after the model call below for the
        # safety net if it's ignored anyway.
        message_parts.append(
            "NOTE: no statute excerpts were retrieved from the Knesset legislation "
            "database for this specific question. If you are not highly confident of "
            "an exact statute, section number, or case citation from well-established "
            "general knowledge, say so explicitly and recommend the user verify "
            "independently or consult a licensed attorney -- do NOT state a specific "
            "citation you cannot verify."
        )
    if file_context:
        message_parts.append(f"The user attached the following document(s):\n\n{file_context}")
    message_parts.append(f"User question: {user_text}")
    combined_message = "\n\n".join(message_parts)

    messages = (
        [{"role": "system", "content": LEGAL_SYSTEM_PROMPT}]
        + clean_history
        + [{"role": "user", "content": combined_message}]
    )
    try:
        answer = _chat_backend(
            messages, model=model, num_predict=num_predict,
            num_ctx=num_ctx, timeout=request_timeout,
            timeout_hint="try a narrower question -- e.g. ask about a specific section rather than a whole law.",
        )
    except RuntimeError as e:
        yield f"⚠️ {e}", gr.skip(), gr.skip()
        return

    answer, citations = _annotate_legal_answer(answer, n_found)
    yield answer, _format_citations_panel(citations), _format_knesset_sources_panel(law_sources, retrieval_info)


def _format_legal_v2_sources_panel(sources: list[dict], verification: dict, cycles_used: int) -> str:
    """Sidebar panel for the DeepSeek-backed Attorney tabs' new
    pipeline -- distinct from _format_knesset_sources_panel (used by the
    Fast tab's older pipeline) since this one also has verification
    status/cycle count to show, and a per-source is_current/effective_from
    flag from the richer metadata schema scripts/embed_local_law_pdfs_bgem3.py
    writes."""
    header = f"**Verification: {verification.get('status', '?')}** ({cycles_used} retrieval cycle(s) used)"
    if verification.get("issues"):
        header += "\n\n" + "\n".join(f"- ⚠️ {issue}" for issue in verification["issues"])
    header += "\n\n---\n\n"

    if not sources:
        return header + _KNESSET_SOURCES_NONE_FOUND

    lines = []
    for s in sources:
        line = f"- **{s.get('law_name', '')}**"
        if s.get("section"):
            line += f" — סעיף {s['section']}"
            if s.get("subsection"):
                line += f"({s['subsection']})"
        if not s.get("is_current", True):
            line += "  \n  🚨 *not currently marked in force*"
        elif s.get("effective_from"):
            line += f"  \n  תחילה: {s['effective_from']}"
        if s.get("source_url"):
            line += f"  \n  [{s['source_url']}]({s['source_url']})"
        lines.append(line)
    return header + "\n\n".join(lines)


def _legal_chat_fn_v2_impl(message, history, hebrew_doc, *, deepseek_model: str, tab_label: str):
    """
    Shared generator behind BOTH DeepSeek-backed Attorney tabs (one
    per DeepSeek size -- see legal_chat_fn / legal_chat_fn_deepseek14b
    below). Every tab-specific choice is passed in explicitly (deepseek_model
    for which Ollama tag runs the reasoning/verification stages, tab_label
    only for the missing-deps message) rather than baked into one function,
    so the two tabs can never silently drift out of sync with each other --
    same reasoning app/ui.py already applies to _legal_chat_fn_impl (the
    OLDER pipeline's shared implementation behind the 24B/1.7B split).

    answer_legal_question() is synchronous and can genuinely take several
    minutes (multiple sequential LLM calls, possibly repeated across retry
    cycles, all on CPU) -- running it directly inside this generator would
    leave the chat UI frozen with no feedback for the whole duration, which
    is exactly the "looks hung" problem this project's other tabs already
    guard against with intermediate status yields. Since
    answer_legal_question reports progress via a plain callback rather than
    being a generator itself, it's run in a background thread here, with
    its on_progress() calls pushed onto a queue that this generator drains
    and yields from as they arrive -- same end result (live progress in the
    chat) as every other tab's yield-per-stage pattern, just bridged across
    a thread boundary since the pipeline module has no Gradio/generator
    dependency of its own (see that module's docstring for why).
    """
    if isinstance(message, dict):
        user_text = message.get("text", "")
        files = message.get("files", []) or []
    else:
        user_text = message
        files = []

    if not LEGAL_V2_AVAILABLE:
        yield (
            f"⚠️ The {tab_label} tab's pipeline (app/legal_pipeline_v2.py) isn't fully set up yet. "
            "Make sure you've run:\n"
            "- `pip install chromadb rank_bm25`\n"
            "- `ollama pull bge-m3`\n"
            f"- `ollama pull {deepseek_model}`\n"
            "- `python scripts/embed_local_law_pdfs_bgem3.py` (builds the `israeli_legal_db` collection)\n\n"
            "Until then, try the **Attorney 1.7B (Fast)** tab instead, which still uses the "
            "original Knesset-RAG pipeline.",
            gr.skip(), gr.skip(),
        )
        return

    file_context = ""
    if files:
        yield "📄 *Extracting text from the attached file(s)…*", gr.skip(), gr.skip()
        file_context = extract_context_from_files(files, hebrew=hebrew_doc)
        if len(file_context) > MAX_CONTEXT_CHARS:
            file_context = file_context[:MAX_CONTEXT_CHARS] + "\n[...truncated, file is longer...]"

    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    progress_queue: queue.Queue = queue.Queue()
    result_holder: dict = {}

    def _run_pipeline():
        try:
            result_holder["result"] = answer_legal_question(
                user_text, clean_history, file_context,
                on_progress=lambda msg: progress_queue.put(("progress", msg)),
                deepseek_model=deepseek_model,
            )
        except RuntimeError as e:
            # Same contract as _chat_backend's RuntimeError elsewhere in
            # this file -- a clean, user-facing message, not a raw traceback.
            result_holder["error"] = str(e)
        except Exception as e:  # noqa: BLE001 -- last-resort net so a bug in
            # the pipeline surfaces as a chat message instead of a silently
            # hung worker thread the generator below would wait on forever.
            result_holder["error"] = f"Unexpected error in the legal pipeline: {e}"
        finally:
            progress_queue.put(("done", None))

    worker = threading.Thread(target=_run_pipeline, daemon=True)
    worker.start()

    while True:
        kind, payload = progress_queue.get()
        if kind == "progress":
            yield payload, gr.skip(), gr.skip()
        else:
            break
    worker.join()

    if "error" in result_holder:
        yield f"⚠️ {result_holder['error']}", gr.skip(), gr.skip()
        return

    result = result_holder["result"]
    citations = _extract_citations(result["answer"])
    yield (
        result["answer"],
        _format_citations_panel(citations),
        _format_legal_v2_sources_panel(result["sources"], result["verification"], result["cycles_used"]),
    )


def legal_chat_fn(message, history, hebrew_doc=False):
    """Attorney 32B tab -- app/legal_pipeline_v2's pipeline (Dicta query
    understanding -> BGE-M3 vector search + BM25 -> Reciprocal Rank Fusion
    -> hybrid rerank -> DeepSeek legal reasoning -> Dicta Hebrew draft ->
    DeepSeek verification, retrying up to MAX_VERIFICATION_CYCLES times on
    FAIL), running DeepSeek-R1-Distill-Qwen-32B for the reasoning and
    verification stages -- stronger legal reasoning, but does not fit in
    RAM alongside DictaLM-3.0-24B-Thinking at the same time on this
    project's target 32GB hardware (see DEEPSEEK_MODEL_DEFAULT's comment in
    app/legal_pipeline_v2.py), so every Dicta<->DeepSeek handoff pays a
    real model-load cost. If that's too slow, use **Attorney 24B (DeepSeek
    14B)** instead."""
    yield from _legal_chat_fn_v2_impl(
        message, history, hebrew_doc,
        deepseek_model=LEGAL_V2_DEEPSEEK_32B, tab_label="Attorney 32B",
    )


def legal_chat_fn_deepseek14b(message, history, hebrew_doc=False):
    """Attorney 24B (DeepSeek 14B) tab -- identical pipeline to legal_chat_fn
    above, just with DeepSeek-R1-Distill-Qwen-14B (~9GB at Q4_K_M) running
    the reasoning/verification stages instead of the 32B variant. That size
    comfortably fits in memory alongside DictaLM-3.0-24B-Thinking on this
    project's target 32GB-RAM/no-GPU hardware, so switching between the
    Dicta and DeepSeek calls each pipeline stage makes doesn't force a
    model reload every time -- meaningfully faster end-to-end, at some cost
    to reasoning/verification depth versus the 32B tab."""
    yield from _legal_chat_fn_v2_impl(
        message, history, hebrew_doc,
        deepseek_model=LEGAL_V2_DEEPSEEK_14B, tab_label="Attorney 24B (Offline AI 14B)",
    )


def legal_chat_fn_fast(message, history, hebrew_doc=False):
    """Attorney 1.7B (Fast) tab -- DictaLM-3.0-1.7B-Thinking. Much faster
    and lower-RAM on the same hardware, at a real cost to legal reasoning
    depth and citation reliability -- see LEGAL_MODEL_FAST's comment.
    Same Knesset RAG grounding and hallucination safety nets as the 24B
    tab, just with a bigger context window (see _LEGAL_FAST_NUM_CTX) since
    a 1.7B model's KV cache is small enough that the 24B tab's RAM-pressure
    problem shouldn't apply here."""
    yield from _legal_chat_fn_impl(
        message, history, hebrew_doc,
        model=LEGAL_MODEL_FAST, num_predict=_LEGAL_FAST_NUM_PREDICT, num_ctx=_LEGAL_FAST_NUM_CTX,
        request_timeout=_LEGAL_FAST_REQUEST_TIMEOUT_SECONDS,
        whole_law_max_weight=_KNESSET_FAST_WHOLE_LAW_MAX_WEIGHT,
    )


# --- Canon AI (RAG over the Codice di Diritto Canonico, Italian) -----------
#
# Unlike the Attorney tabs (which stuff an attached document straight into
# the prompt), this tab retrieves relevant canons from a local ChromaDB
# vector store -- built ahead of time by scrape_cic_it.py + embed_to_chroma.py
# -- and grounds the model's answer in whatever it actually retrieves.
#
# Path/collection name here MUST match what embed_to_chroma.py was run
# with (--chroma-dir / --collection). Adjust if you used different values.
CANON_CHROMA_DIR = str(Path(__file__).resolve().parent / "chroma_db")
CANON_COLLECTION_NAME = "cic_it"

# nomic-embed-text via Ollama -- same model/endpoint used to build the
# collection. Query-time text needs the "search_query: " prefix (NOT
# "search_document: ", which is what was used when indexing) -- nomic's
# model was trained with different prefixes for each side of retrieval,
# and mixing them up silently degrades results rather than erroring.
_CANON_OLLAMA_EMBED_URL = "http://localhost:11434/api/embeddings"
_CANON_EMBED_MODEL = "nomic-embed-text"
_CANON_QUERY_PREFIX = "search_query: "

# How many canons to retrieve per question. Higher = more coverage but
# more context spent on possibly-irrelevant canons.
# Raised 5 -> 8: with only 5 slots, a relevant-but-not-top-1 canon pulled
# in by _CANON_FETCH_K's wider net was still frequently getting trimmed
# off again by the re-rank step below before it ever reached the LLM.
# Re-validate with test_canon_rag.py (Recall@1/@3/@5, MRR) if you tune
# this further -- there's a real prompt-size/latency cost to going higher,
# especially now that generation runs on the larger gpt-oss:20b model.
_CANON_TOP_K = 8

# Model that answers using the retrieved canons. Pinned explicitly to
# gpt-oss:20b (same model the plain Chat tab uses) rather than left as None
# to inherit the backend's OLLAMA_MODEL default -- Canon AI and Chat are
# deliberately meant to share this model, but making it explicit here means
# that stays true even if OLLAMA_MODEL is ever repurposed/changed in
# main.py for something else (e.g. invoice classification) without this
# tab silently following along. Set to a different Ollama model string
# (pulled and available to the backend) if you want Canon AI on its own
# dedicated model instead, e.g. one of the LEGAL_MODEL constants above.
CANON_MODEL = "gpt-oss:20b"
_CANON_NUM_PREDICT = 2048
_CANON_REQUEST_TIMEOUT_SECONDS = 900

CANON_SYSTEM_PROMPT = (
    "You are a canon lawyer -- an attorney specialized in the Roman "
    "Catholic Church's Codice di Diritto Canonico (Code of Canon Law). "
    "Think through and answer every question strictly according to the "
    "canons provided to you below, retrieved from the official Italian "
    "text on vatican.va -- not from general recall of canon law, and "
    "not from the law of any civil jurisdiction, unless the user "
    "explicitly asks about civil law.\n\n"
    "For every substantive claim, cite the specific canon(s) you are "
    "relying on immediately after the claim -- for example: 'A parish "
    "priest is removed only for a grave cause (Can. 1740).' Base your "
    "answer ONLY on the canons retrieved below plus general background "
    "knowledge of canon law's structure -- never invent a canon number "
    "or a rule that isn't in the excerpts you were given. If the "
    "retrieved canons don't actually answer the question, say so "
    "explicitly instead of guessing -- a wrong or fabricated citation is "
    "worse than admitting the retrieval didn't cover it.\n\n"
    "Keep your answer proportionate to the question: for a broad or "
    "general topic, cover the most important, directly relevant canons "
    "rather than exhaustively enumerating every retrieved excerpt -- you "
    "can offer to go deeper on a specific canon if the user wants that.\n\n"
    "Note where relevant that Book VI (Cann. 1311-1399) was fully "
    "reformed in 2021 ('Pascite Gregem Dei') -- flag if a retrieved Book "
    "VI canon might reflect the pre-2021 text.\n\n"
    "You are not a substitute for advice from a canon lawyer engaged by "
    "the person's diocese or tribunal, or for the guidance of competent "
    "Church authority, and you should say so when a question calls for "
    "one. Always reply in the same language the user's question is "
    "written in -- e.g. if they write in English, your entire answer "
    "must be in English, even though the retrieved canon text is in "
    "Italian. Never switch languages on the user unasked."
)

_CANON_SOURCES_PLACEHOLDER = (
    "_No canons retrieved yet. Once you ask a question, the specific "
    "canons retrieved from the vector database will be listed here._"
)
_CANON_SOURCES_NONE_FOUND = (
    "_Nothing was retrieved for this question -- the answer above (if any) "
    "isn't grounded in a specific canon. Try rephrasing._"
)
_CANON_DB_MISSING_MSG = (
    "⚠️ Canon AI's vector database isn't available. Make sure you've run "
    "scrape_cic_it.py then embed_to_chroma.py (pointed at "
    f"'{CANON_CHROMA_DIR}', collection '{CANON_COLLECTION_NAME}'), and "
    "that chromadb is installed (`pip install chromadb`)."
)

_canon_collection_cache = None


def _get_canon_collection():
    """Lazily connect to the persistent Chroma collection, caching the
    handle across calls. Returns None (rather than raising) on any
    failure -- callers surface a clean in-chat error instead of crashing
    the whole UI process if the vector DB hasn't been built yet."""
    global _canon_collection_cache
    if _canon_collection_cache is not None:
        return _canon_collection_cache
    if chromadb is None:
        return None
    try:
        client = chromadb.PersistentClient(path=CANON_CHROMA_DIR)
        _canon_collection_cache = client.get_collection(CANON_COLLECTION_NAME)
        return _canon_collection_cache
    except Exception as e:
        import traceback
        print(f"[Canon AI] failed to open Chroma collection at "
              f"'{CANON_CHROMA_DIR}' / '{CANON_COLLECTION_NAME}': {e}")
        traceback.print_exc()
        return None


def _embed_canon_query(text: str) -> list[float] | None:
    try:
        resp = requests.post(
            _CANON_OLLAMA_EMBED_URL,
            json={"model": _CANON_EMBED_MODEL, "prompt": _CANON_QUERY_PREFIX + text},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["embedding"]
    except Exception as e:
        # Previously silent (bare `except Exception: return None`) -- this was
        # genuinely undiagnosable: a caller saw only a generic "database
        # unavailable" message with NOTHING in the terminal to distinguish
        # it from a real ChromaDB connection failure (which _get_*_collection
        # already logs). Most common real cause: Ollama evicting this model
        # to load a different one under OLLAMA_MAX_LOADED_MODELS=1 (see
        # run.ps1) and the 30s timeout above hitting mid-swap.
        print(f"[RAG embed] embedding call to Ollama ({_CANON_EMBED_MODEL}) failed: {e}")
        return None


# Matches how a canon number is named across English, Latin-abbreviated,
# and Italian phrasing: "can. 1055" / "canon 1055" / "canone 1055" /
# "canoni 1055" / "cann. 1055" / "can. n. 1055". Italian users naturally
# write "canone"/"canoni" (not the Latin/English "can./canon" the previous
# version of this regex only covered), which silently meant the
# exact-match lookup below never fired for them at all.
_CANON_NUMBER_RE = re.compile(
    r"\bcann?(?:on(?:e|i)?|s)?\.?\s*(?:n(?:um)?\.?\s*)?(\d{1,4})\b",
    re.IGNORECASE,
)

# How many candidates to pull from the vector index before re-ranking down
# to _CANON_TOP_K. Over-fetching gives the re-ranker real material to work
# with -- with n_results=5 the LLM never even sees a 6th-closest canon that
# might actually have been the better answer; casting a wider net first and
# then re-ranking recovers those cases.
# Raised 15 -> 25 alongside the _CANON_TOP_K increase above: the true best
# match for a semantic (non-numeric) query isn't always in Chroma's raw
# top 15 by pure vector distance -- widening the pool before hybrid
# re-ranking reduces the chance it gets excluded before it's even scored.
_CANON_FETCH_K = 25

# Weight given to lexical (keyword) overlap vs. vector similarity when
# re-ranking. Pure vector similarity can rank a topically-similar-but-wrong
# canon above one that shares the query's actual legal terms; blending in
# lexical overlap corrects for that without needing a cross-encoder model.
_CANON_LEXICAL_WEIGHT = 0.35

# Absolute floor on cosine similarity (see _cosine_similarity below) under
# which a candidate is treated as NOT relevant at all -- dropped outright,
# not just ranked lower. Shared across every tab that calls
# _rerank_candidates (Canon AI / GDPR AI / HIPAA AI / Knesset RAG), since
# they all had the same underlying bug: Chroma's collection.query() always
# returns its n nearest neighbours, no matter how far away every single one
# of them actually is. Asking about a law/canon/article that was never
# embedded still silently returned the "least-bad" match dressed up as a
# real hit, because the old scoring (vector_sim = 1 - dist/max_dist) was
# normalized against the OTHER CANDIDATES IN THE SAME BATCH, not against
# any absolute notion of "close enough to be relevant" -- that guarantees
# the single least-irrelevant candidate always comes out looking like a
# near-perfect match. Confirmed directly in practice: asking the Attorney
# tab about the Clean Air Law, with only Privacy Law embedded in the
# database, still returned "relevant" statute excerpts and a
# grounded-sounding citation.
#
# 0.45 (the original value here) turned out to be too strict in practice --
# confirmed directly: a Canon AI question about a topic that genuinely IS in
# the vector database still came back "no canons matching". Lowered to 0.20
# as a much safer starting point.
#
# Two things were almost certainly stacking against the original 0.45:
#   1. It was a guess, not a calibrated number (same caveat this file
#      already gives _CANON_LEXICAL_WEIGHT/_CANON_TOP_K etc.).
#   2. nomic-embed-text-v1.5 (confirmed as the model in use, from the
#      Ollama server log) is a primarily ENGLISH-trained embedding model,
#      not a multilingual one (unlike e.g. multilingual-e5 or LaBSE). Canon
#      AI is explicitly designed to let you ask in English against
#      Italian-indexed canon text (see CANON_SYSTEM_PROMPT) -- but an
#      English query embedded against an Italian document embedding likely
#      has a STRUCTURALLY lower cosine similarity than a same-language
#      pair would, purely because this model was never trained to align
#      languages well, not because the match is actually weaker. A single
#      flat threshold tuned assuming same-language matching will reject
#      genuinely relevant cross-lingual hits. This applies to the Knesset
#      DB too if questions are asked in English against Hebrew statute
#      text -- worth checking whether Hebrew/Italian corpora need a lower
#      floor than an all-English corpus (like GDPR/HIPAA) once you have
#      real numbers from each tab.
#
# _rerank_candidates now prints the FULL spread of candidate similarities
# on every query (not just the best one), specifically so a single log
# paste shows exactly where real matches land for a given corpus/language
# pair, instead of tuning this blind one guess at a time. Raise it if
# genuinely off-topic questions still come back "grounded"; lower it
# further if a question you know is covered still comes back as "nothing
# found".
_MIN_RELEVANCE_SIMILARITY = 0.20

_STOPWORDS = frozenset("""
a an the of to in on for and or is are was were be been being this that
these those what which who whom does do did can may must shall not with
as by from about into over under between it its it's i we you he she they
""".split() + """
il lo la i gli le un uno una del dello della dei degli delle al allo alla
ai agli alle dal dallo dalla dai dagli dalle nel nello nella nei negli
nelle sul sullo sulla sui sugli sulle col coi che chi cosa come dove
quando perche perché non ma se anche piu più meno molto questo questa
questi queste quello quella quelli quelle ed cioe cioè dice dicono essere
sono stato stata essendo fra tra con per su da di
""".split())


def _strip_accents(text: str) -> str:
    """NFKD-decomposes accented characters and drops the combining marks
    (e.g. 'nullità' -> 'nullita'), so Italian words match consistently
    without the previous behaviour of the plain [a-z'] regex silently
    truncating a word at its first accented character (e.g. 'nullit')."""
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def _tokenize(text: str) -> set[str]:
    ascii_text = _strip_accents(text.lower())
    return {w for w in re.findall(r"[a-z']+", ascii_text) if w not in _STOPWORDS and len(w) > 2}


def _has_meaningful_content(query: str) -> bool:
    """
    True if the query tokenizes to at least one real (non-stopword,
    length > 2) term. Used to skip semantic/vector search entirely for
    greetings and chit-chat ("hi", "hello", "thanks") -- a DIFFERENT
    failure mode than the one _MIN_RELEVANCE_SIMILARITY guards against.
    That floor assumes the query has a real topic which simply isn't
    covered by anything embedded; it does nothing for a query that has no
    real topic at all. Sentence embeddings are anisotropic (real-world
    embeddings cluster in a narrow cone of the vector space rather than
    spreading out evenly), so a short, content-free query's embedding can
    still land deceptively close -- by raw cosine similarity -- to some
    arbitrary document, purely as an artifact of the embedding space, not
    because of any genuine topical relevance. Confirmed directly in
    practice: asking Canon AI a plain "hi" still populated the "Canons
    retrieved" sidebar. Cheaper and more reliable to just not run the
    embedding call/vector search at all below this bar -- exact-citation
    lookups (a canon/article/section number) are unaffected, since those
    run through their own regex match, not this check.
    """
    return bool(_tokenize(query))


def _lexical_overlap_score(query_tokens: set[str], doc_text: str) -> float:
    """Fraction of the query's meaningful terms that also appear in the
    candidate canon's text -- a cheap stand-in for BM25 that needs no
    extra dependencies or index beyond what's already in memory."""
    if not query_tokens:
        return 0.0
    doc_tokens = _tokenize(doc_text)
    return len(query_tokens & doc_tokens) / len(query_tokens)


def _embeddings_to_list(value):
    """Chroma's query()/get() results return `embeddings` as a numpy array
    in this chromadb version -- unlike documents/metadatas/distances, which
    all come back as plain Python lists/dicts here. A plain
    `value or [[]]` truthiness check (the pattern used elsewhere in this
    file for documents/metadatas/distances, which works fine for those)
    crashes on a numpy array instead of behaving like a normal Python
    falsy/truthy check: numpy raises "The truth value of an array with
    more than one element is ambiguous. Use a.any() or a.all()" rather than
    silently picking one.

    A single top-level .tolist() call is NOT always enough to fully fix
    this: if Chroma returns embeddings as a numpy array of dtype=object
    (each row itself a separate numpy array, rather than one uniform
    stacked N-D array -- confirmed to be happening here, since the error
    this function exists to prevent recurred even after the first,
    shallower version of this fix), .tolist() only unwraps the OUTER
    object array. Each inner row stays a real numpy array underneath,
    which then still crashes the exact same way the first time anything
    downstream (e.g. _cosine_similarity) does a plain truthiness check on
    it. This recurses until every leaf is a genuine Python float/int, no
    matter how many layers of numpy-ness are nested inside."""
    if hasattr(value, "tolist"):
        value = value.tolist()
    if isinstance(value, list):
        return [_embeddings_to_list(v) for v in value]
    return value


def _cosine_similarity(a, b) -> float:
    """True cosine similarity between two raw embedding vectors, computed
    directly rather than derived from Chroma's returned `distances`. This
    is what makes _MIN_RELEVANCE_SIMILARITY a meaningful absolute
    threshold: this collection was created via a bare
    get_or_create_collection() call with no explicit hnsw:space metadata,
    so whether Chroma is internally using l2 or cosine distance isn't
    something this code controls or can safely assume -- computing
    similarity ourselves from the vectors sidesteps that question
    entirely, and (just as important) gives a score that's comparable
    ACROSS different queries, unlike a distance normalized against
    whatever else happened to be fetched alongside it in one batch.

    Deliberately does NOT do `if not a or not b:` here -- that's the exact
    line that crashed with numpy's "ambiguous truth value" error even
    after _embeddings_to_list was added at the call sites, because a
    caller passing in a still-numpy value (however that happens) hits the
    same bare truthiness check a second time. len(...) == 0 works
    identically for a plain Python list AND a numpy array without ever
    evaluating the array's overall truthiness, so this is safe regardless
    of what type a/b actually are -- a deliberate second layer of defense
    on top of _embeddings_to_list, not a replacement for it."""
    if a is None or b is None or len(a) == 0 or len(b) == 0:
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = sum(x * x for x in a) ** 0.5
    norm_b = sum(x * x for x in b) ** 0.5
    if norm_a == 0.0 or norm_b == 0.0:
        return 0.0
    return dot / (norm_a * norm_b)


def _rerank_candidates(
    query: str, query_vector: list[float], docs: list[str], metas: list[dict],
    embeddings: list[list[float]], top_n: int, min_similarity: float,
    label: str = "",
):
    """
    Over-fetch -> hybrid re-score (vector similarity + lexical overlap) ->
    absolute relevance floor -> lightweight diversity filter, replacing
    plain top-K vector similarity. Returns (docs, metas) trimmed to top_n,
    most relevant and least redundant first -- or ([], []) if NOTHING
    clears min_similarity, which is the actual anti-hallucination fix: see
    _MIN_RELEVANCE_SIMILARITY's comment for the real failure mode this
    closes (a query about content that was never embedded still used to
    come back "grounded" in whatever was closest, however irrelevant).

    A candidate below min_similarity is dropped outright, not merely
    down-ranked. If every candidate is dropped, downstream code already
    treats an empty result exactly like "nothing was retrieved" (see each
    retrieve_* function's docstring) -- the model gets told plainly that
    nothing matched, instead of being handed irrelevant context dressed up
    as a real citation.
    """
    if not docs:
        return [], []

    query_tokens = _tokenize(query)
    scored = []
    similarities = []
    for doc, meta, emb in zip(docs, metas, embeddings):
        vector_sim = _cosine_similarity(query_vector, emb)
        similarities.append(vector_sim)
        if vector_sim < min_similarity:
            continue
        lexical = _lexical_overlap_score(query_tokens, doc)
        hybrid = (1 - _CANON_LEXICAL_WEIGHT) * vector_sim + _CANON_LEXICAL_WEIGHT * lexical
        scored.append((hybrid, doc, meta))

    best_similarity = max(similarities) if similarities else -1.0
    # Prints the top handful of raw similarities on EVERY query (not just
    # failures) so tuning _MIN_RELEVANCE_SIMILARITY needs one log paste,
    # not repeated blind guesses -- shows both where real matches land
    # (when something clears the floor) and how close a rejected query
    # actually came (when nothing does).
    top_sims_str = ", ".join(f"{s:.3f}" for s in sorted(similarities, reverse=True)[:8])
    tag = f"[{label}] " if label else ""
    if not scored:
        print(f"{tag}rerank: best candidate similarity {best_similarity:.3f} never "
              f"cleared the {min_similarity:.2f} floor across {len(docs)} candidate(s) "
              f"-- treating as nothing relevant found. Top similarities: [{top_sims_str}]")
        return [], []
    print(f"{tag}rerank: best candidate similarity {best_similarity:.3f}, "
          f"{len(scored)}/{len(docs)} candidate(s) cleared the {min_similarity:.2f} floor. "
          f"Top similarities: [{top_sims_str}]")
    scored.sort(key=lambda x: x[0], reverse=True)

    # Diversity filter: greedily keep the next-best candidate only if it
    # doesn't overlap too heavily (by shared tokens) with something already
    # selected, so near-duplicate/adjacent canons don't crowd out the top_n
    # slots at the expense of covering the question from other angles.
    selected_docs, selected_metas, selected_tokens = [], [], []
    for hybrid, doc, meta in scored:
        if len(selected_docs) >= top_n:
            break
        doc_tokens = _tokenize(doc)
        # Raised 0.75 -> 0.85: two canons on the same topic can legitimately
        # share a lot of vocabulary (e.g. adjacent canons in the same
        # article) while still being distinct, correct answers -- 0.75 was
        # sometimes excluding a genuinely relevant canon just for being
        # lexically close to one already selected. Re-check with
        # test_canon_rag.py if you tune this further; too high and
        # near-duplicate/redundant canons start crowding out real coverage
        # again.
        too_similar = any(
            len(doc_tokens & prev) / max(1, min(len(doc_tokens), len(prev))) > 0.85
            for prev in selected_tokens
        )
        if too_similar and len(selected_docs) > 0:
            continue
        selected_docs.append(doc)
        selected_metas.append(meta)
        selected_tokens.append(doc_tokens)

    return selected_docs, selected_metas


def retrieve_canons(query: str, n_results: int = _CANON_TOP_K):
    """
    Retrieves relevant canons for a query, combining three strategies:

    1. Exact match: if the query names a specific canon number (e.g. "canon
       1055" or "can. 129"), that canon is fetched directly via a metadata
       filter and placed first. Vector similarity alone is unreliable for
       exact lookups -- a query embedding for "canon 1055" isn't guaranteed
       to land closest to canon 1055's own embedding, especially for canons
       with short or generic text.
    2. Semantic search + hybrid re-rank: the query is embedded with
       nomic-embed-text, _CANON_FETCH_K candidates are pulled by vector
       similarity, then re-scored by blending vector similarity with
       lexical keyword overlap and filtered for diversity, before trimming
       down to n_results. This catches cases where the plain nearest
       neighbour isn't actually the best answer, without needing a
       cross-encoder model.
    3. Diversity filtering (part of the re-rank step above) avoids near-
       duplicate/adjacent canons crowding out the final n_results slots.

    Returns (context_text, sources) where sources is a list of dicts used
    to build the sidebar panel -- or (None, None) on any failure (missing
    DB, unreachable Ollama, etc.), which callers turn into a clear in-chat
    message rather than a silent empty answer.
    """
    collection = _get_canon_collection()
    if collection is None:
        return None, None

    exact_docs, exact_metas = [], []
    for canon_num in dict.fromkeys(_CANON_NUMBER_RE.findall(query)):  # dedupe, keep order
        for candidate in (canon_num, int(canon_num)):  # metadata may be str or int
            try:
                hit = collection.get(where={"canon_number": candidate})
            except Exception:
                continue
            for doc, meta in zip(hit.get("documents") or [], hit.get("metadatas") or []):
                exact_docs.append(doc)
                exact_metas.append(meta)
            if hit.get("documents"):
                break  # found it under this type, no need to try the other

    query_vector = None
    if _has_meaningful_content(query):
        query_vector = _embed_canon_query(query)
        if query_vector is None and not exact_docs:
            return None, None
    elif not exact_docs:
        return "", []  # greeting/chit-chat, no real topic to search on and no exact citation either

    vector_docs, vector_metas = [], []
    if query_vector is not None:
        try:
            results = collection.query(
                query_embeddings=[query_vector],
                n_results=_CANON_FETCH_K,
                include=["documents", "metadatas", "distances", "embeddings"],
            )
            raw_docs = (results.get("documents") or [[]])[0]
            raw_metas = (results.get("metadatas") or [[]])[0]
            raw_dists = (results.get("distances") or [[]])[0]
            raw_embeds = (_embeddings_to_list(results.get("embeddings")) or [[]])[0]
            vector_docs, vector_metas = _rerank_candidates(
                query, query_vector, raw_docs, raw_metas, raw_embeds,
                top_n=n_results, min_similarity=_MIN_RELEVANCE_SIMILARITY,
                label="Canon AI",
            )
        except Exception:
            pass  # exact matches (if any) still make it through below

    # Merge exact matches first, then vector results, de-duping by canon
    # number so a canon named explicitly doesn't also show up twice.
    seen = set()
    documents, metadatas = [], []
    for doc, meta in zip(exact_docs + vector_docs, exact_metas + vector_metas):
        key = meta.get("canon_number", "?")
        if key in seen:
            continue
        seen.add(key)
        documents.append(doc)
        metadatas.append(meta)

    if not documents:
        return "", []

    context_parts = []
    sources = []
    for doc, meta in zip(documents, metadatas):
        canon_num = meta.get("canon_number", "?")
        path = meta.get("hierarchy_path", "")
        url = meta.get("source_url", "")
        note = meta.get("in_force_note", "")
        context_parts.append(f"[Can. {canon_num}] ({path})\n{doc}")
        sources.append({"canon_number": canon_num, "hierarchy_path": path, "source_url": url, "note": note})

    context_text = "\n\n---\n\n".join(context_parts)
    return context_text, sources


def _format_canon_sources_panel(sources: list[dict] | None) -> str:
    if not sources:
        return _CANON_SOURCES_NONE_FOUND
    lines = []
    for s in sources:
        line = f"- **Can. {s['canon_number']}** — {s['hierarchy_path']}"
        if s.get("source_url"):
            line += f"  \n  [{s['source_url']}]({s['source_url']})"
        if s.get("note"):
            line += f"  \n  ⚠️ {s['note']}"
        lines.append(line)
    return "\n\n".join(lines)


def canon_chat_fn(message, history):
    """
    RAG chat over the scraped/embedded Codice di Diritto Canonico. Not
    multimodal (no file attachment) -- the whole point of this tab is
    grounding answers in the pre-built vector database rather than
    whatever the user happens to attach.

    Returns (answer, sources_panel_markdown) -- see canon_sources_panel's
    wiring below, same additional_outputs pattern as the Attorney tabs'
    citations panel.
    """
    user_text = message.get("text", message) if isinstance(message, dict) else message

    if chromadb is None or _get_canon_collection() is None:
        yield _CANON_DB_MISSING_MSG, gr.skip()
        return

    yield "🔍 *Searching ChromaDB for canons relevant to your question…*", gr.skip()

    context_text, sources = retrieve_canons(user_text)
    if context_text is None:
        yield _CANON_DB_MISSING_MSG, gr.skip()
        return

    n_found = len(sources) if sources else 0
    if n_found:
        yield (
            f"📖 *Found {n_found} relevant canon(s) — asking the model to "
            f"answer using them…*",
            gr.skip(),
        )
    else:
        yield "🤔 *Nothing matched in the vector database — asking the model how to proceed…*", gr.skip()

    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    if context_text:
        combined_message = (
            f"Retrieved canons relevant to the question:\n\n{context_text}\n\n"
            f"User question: {user_text}"
        )
    else:
        combined_message = (
            "No canons were retrieved for this question -- tell the user "
            f"that plainly instead of guessing.\n\nUser question: {user_text}"
        )

    messages = (
        [{"role": "system", "content": CANON_SYSTEM_PROMPT}]
        + clean_history
        + [{"role": "user", "content": combined_message}]
    )
    try:
        answer = _chat_backend(
            messages, model=CANON_MODEL, num_predict=_CANON_NUM_PREDICT,
            timeout=_CANON_REQUEST_TIMEOUT_SECONDS,
            timeout_hint="try a narrower question about a specific canon or topic.",
        )
    except RuntimeError as e:
        yield f"⚠️ {e}", gr.skip()
        return

    yield answer, _format_canon_sources_panel(sources)


# --- GDPR AI (RAG over the GDPR -- Regulation (EU) 2016/679) ---------------
#
# Same shape as Canon AI just above (retrieve -> ground -> answer over a
# pre-built local vector store), pointed at a different corpus: the GDPR's
# 99 articles, scraped article-by-article via scripts/scrape_gdpr.py and
# embedded via scripts/embed_gdpr_to_chroma.py. Deliberately reuses Canon
# AI's embedding/re-ranking machinery rather than duplicating it --
# _embed_canon_query, _rerank_candidates, _tokenize, _lexical_overlap_score,
# and _strip_accents are all already fully generic (nomic-embed-text via
# Ollama, hybrid vector+lexical scoring with a diversity filter): nothing
# about them is actually canon-specific despite the name, so the GDPR
# collection below just calls straight into them.
#
# Path/collection name here MUST match what embed_gdpr_to_chroma.py was run
# with (--chroma-dir / --collection). Adjust if you used different values.
GDPR_CHROMA_DIR = str(Path(__file__).resolve().parent / "chroma_db")
GDPR_COLLECTION_NAME = "gdpr"

# How many articles to retrieve per question -- same reasoning/tuning basis
# as _CANON_TOP_K/_CANON_FETCH_K (see those comments): wide enough that a
# relevant-but-not-top-1 article isn't trimmed off before the LLM ever sees
# it, without spending so much context on marginal matches that latency/
# prompt size suffers. Re-tune per corpus size if this ever grows past the
# GDPR's fixed 99 articles (e.g. if implementing acts/case law are added).
_GDPR_TOP_K = 6
_GDPR_FETCH_K = 20

# Model that answers using the retrieved articles. Pinned explicitly to
# gpt-oss:20b (same model backing Chat and Canon AI) for the same reason
# CANON_MODEL is pinned rather than left to inherit main.py's OLLAMA_MODEL
# default: GDPR AI is meant to share this model deliberately, and staying
# explicit here means that stays true even if OLLAMA_MODEL is later
# repurposed in main.py for something unrelated (e.g. invoice classification).
GDPR_MODEL = "gpt-oss:20b"
_GDPR_NUM_PREDICT = 2048
_GDPR_REQUEST_TIMEOUT_SECONDS = 900

GDPR_SYSTEM_PROMPT = (
    "You are a data protection / privacy law expert specializing in the EU "
    "General Data Protection Regulation (Regulation (EU) 2016/679, the "
    "'GDPR'). Think through and answer every question strictly according "
    "to the GDPR articles provided to you below -- not from general recall "
    "of privacy law, and not from the law of any other jurisdiction (e.g. "
    "CCPA, UK GDPR post-Brexit divergence, national implementing laws), "
    "unless the user explicitly asks about that instead.\n\n"
    "For every substantive claim, cite the specific article (and "
    "paragraph/point where relevant) you are relying on immediately after "
    "the claim -- for example: 'Processing must have a lawful basis such "
    "as consent or legitimate interest (Art. 6(1) GDPR).' Base your answer "
    "ONLY on the articles retrieved below plus general background "
    "knowledge of the GDPR's structure -- never invent an article number "
    "or a rule that isn't in the excerpts you were given. If the retrieved "
    "articles don't actually answer the question, say so explicitly "
    "instead of guessing -- a wrong or fabricated citation is worse than "
    "admitting the retrieval didn't cover it.\n\n"
    "Keep your answer proportionate to the question: for a broad or "
    "general topic, cover the most important, directly relevant articles "
    "rather than exhaustively enumerating every retrieved excerpt -- you "
    "can offer to go deeper on a specific article if the user wants that.\n\n"
    "You are not a substitute for advice from a qualified data protection "
    "lawyer or a Data Protection Officer, and you should say so when a "
    "question calls for one. Always reply in the same language the user's "
    "question is written in. Never switch languages on the user unasked."
)

_GDPR_SOURCES_PLACEHOLDER = (
    "_No articles retrieved yet. Once you ask a question, the specific "
    "GDPR articles retrieved from the vector database will be listed here._"
)
_GDPR_SOURCES_NONE_FOUND = (
    "_Nothing was retrieved for this question -- the answer above (if any) "
    "isn't grounded in a specific article. Try rephrasing._"
)
_GDPR_DB_MISSING_MSG = (
    "⚠️ GDPR AI's vector database isn't available. Make sure you've run "
    "scripts/scrape_gdpr.py then scripts/embed_gdpr_to_chroma.py (pointed "
    f"at '{GDPR_CHROMA_DIR}', collection '{GDPR_COLLECTION_NAME}'), and "
    "that chromadb is installed (`pip install chromadb`)."
)

_gdpr_collection_cache = None


def _get_gdpr_collection():
    """Lazily connect to the persistent Chroma collection, caching the
    handle across calls. Returns None (rather than raising) on any
    failure -- callers surface a clean in-chat error instead of crashing
    the whole UI process if the vector DB hasn't been built yet."""
    global _gdpr_collection_cache
    if _gdpr_collection_cache is not None:
        return _gdpr_collection_cache
    if chromadb is None:
        return None
    try:
        client = chromadb.PersistentClient(path=GDPR_CHROMA_DIR)
        _gdpr_collection_cache = client.get_collection(GDPR_COLLECTION_NAME)
        return _gdpr_collection_cache
    except Exception as e:
        import traceback
        print(f"[GDPR AI] failed to open Chroma collection at "
              f"'{GDPR_CHROMA_DIR}' / '{GDPR_COLLECTION_NAME}': {e}")
        traceback.print_exc()
        return None


# Matches how an article is named in a question: "Article 6", "Art. 6",
# "art 6", "GDPR Art 6", "article 6(1)(f)" -- captures just the leading
# article number; any paragraph/point suffix like "(1)(f)" is ignored for
# lookup purposes (retrieval is per-article, not per-sub-paragraph) but the
# user's phrasing still matches fine since the number itself is what's
# extracted.
_GDPR_ARTICLE_NUMBER_RE = re.compile(r"\bart(?:icles?)?\.?\s*(\d{1,3})\b", re.IGNORECASE)


def retrieve_gdpr(query: str, n_results: int = _GDPR_TOP_K):
    """
    Retrieves relevant GDPR articles for a query, combining two strategies
    (same overall approach as retrieve_canons -- see that function's
    docstring for the full reasoning):

    1. Exact match: if the query names a specific article number (e.g.
       "Article 6" or "art. 17"), that article is fetched directly via a
       metadata filter and placed first. Vector similarity alone is
       unreliable for exact lookups.
    2. Semantic search + hybrid re-rank: the query is embedded with
       nomic-embed-text (via the shared _embed_canon_query helper),
       _GDPR_FETCH_K candidates are pulled by vector similarity, then
       re-scored by blending vector similarity with lexical keyword
       overlap and filtered for diversity (via the shared
       _rerank_candidates helper), before trimming down to n_results.

    Returns (context_text, sources) where sources is a list of dicts used
    to build the sidebar panel -- or (None, None) on any failure (missing
    DB, unreachable Ollama, etc.), which callers turn into a clear in-chat
    message rather than a silent empty answer.
    """
    collection = _get_gdpr_collection()
    if collection is None:
        return None, None

    exact_docs, exact_metas = [], []
    for art_num in dict.fromkeys(_GDPR_ARTICLE_NUMBER_RE.findall(query)):  # dedupe, keep order
        for candidate in (art_num, int(art_num)):  # metadata may be str or int
            try:
                hit = collection.get(where={"article_number": candidate})
            except Exception:
                continue
            for doc, meta in zip(hit.get("documents") or [], hit.get("metadatas") or []):
                exact_docs.append(doc)
                exact_metas.append(meta)
            if hit.get("documents"):
                break  # found it under this type, no need to try the other

    query_vector = None
    if _has_meaningful_content(query):
        query_vector = _embed_canon_query(query)  # generic nomic-embed-text helper, reused as-is
        if query_vector is None and not exact_docs:
            return None, None
    elif not exact_docs:
        return "", []  # greeting/chit-chat, no real topic to search on and no exact citation either

    vector_docs, vector_metas = [], []
    if query_vector is not None:
        try:
            results = collection.query(
                query_embeddings=[query_vector],
                n_results=_GDPR_FETCH_K,
                include=["documents", "metadatas", "distances", "embeddings"],
            )
            raw_docs = (results.get("documents") or [[]])[0]
            raw_metas = (results.get("metadatas") or [[]])[0]
            raw_dists = (results.get("distances") or [[]])[0]
            raw_embeds = (_embeddings_to_list(results.get("embeddings")) or [[]])[0]
            vector_docs, vector_metas = _rerank_candidates(
                query, query_vector, raw_docs, raw_metas, raw_embeds,
                top_n=n_results, min_similarity=_MIN_RELEVANCE_SIMILARITY,
                label="GDPR AI",
            )
        except Exception:
            pass  # exact matches (if any) still make it through below

    # Merge exact matches first, then vector results, de-duping by article
    # number so an article named explicitly doesn't also show up twice.
    seen = set()
    documents, metadatas = [], []
    for doc, meta in zip(exact_docs + vector_docs, exact_metas + vector_metas):
        key = meta.get("article_number", "?")
        if key in seen:
            continue
        seen.add(key)
        documents.append(doc)
        metadatas.append(meta)

    if not documents:
        return "", []

    context_parts = []
    sources = []
    for doc, meta in zip(documents, metadatas):
        art_num = meta.get("article_number", "?")
        title = meta.get("title", "")
        path = meta.get("hierarchy_path", "")
        url = meta.get("source_url", "")
        recitals = meta.get("recitals", "")
        context_parts.append(f"[Art. {art_num} -- {title}] ({path})\n{doc}")
        sources.append({
            "article_number": art_num, "title": title, "hierarchy_path": path,
            "source_url": url, "recitals": recitals,
        })

    context_text = "\n\n---\n\n".join(context_parts)
    return context_text, sources


def _format_gdpr_sources_panel(sources: list[dict] | None) -> str:
    if not sources:
        return _GDPR_SOURCES_NONE_FOUND
    lines = []
    for s in sources:
        line = f"- **Art. {s['article_number']}** — {s['title']}"
        if s.get("hierarchy_path"):
            line += f"  \n  {s['hierarchy_path']}"
        if s.get("source_url"):
            line += f"  \n  [{s['source_url']}]({s['source_url']})"
        if s.get("recitals"):
            line += f"  \n  📎 Recitals: {s['recitals']}"
        lines.append(line)
    return "\n\n".join(lines)


def gdpr_chat_fn(message, history):
    """
    RAG chat over the scraped/embedded GDPR. Not multimodal (no file
    attachment) -- same reasoning as canon_chat_fn: the point of this tab
    is grounding answers in the pre-built vector database, not whatever
    the user happens to attach.

    Returns (answer, sources_panel_markdown) -- see gdpr_sources_panel's
    wiring below, same additional_outputs pattern as Canon AI / the
    Attorney tabs' citations panel.
    """
    user_text = message.get("text", message) if isinstance(message, dict) else message

    if chromadb is None or _get_gdpr_collection() is None:
        yield _GDPR_DB_MISSING_MSG, gr.skip()
        return

    yield "🔍 *Searching ChromaDB for GDPR articles relevant to your question…*", gr.skip()

    context_text, sources = retrieve_gdpr(user_text)
    if context_text is None:
        yield _GDPR_DB_MISSING_MSG, gr.skip()
        return

    n_found = len(sources) if sources else 0
    if n_found:
        yield (
            f"📜 *Found {n_found} relevant article(s) — asking the model to "
            f"answer using them…*",
            gr.skip(),
        )
    else:
        yield "🤔 *Nothing matched in the vector database — asking the model how to proceed…*", gr.skip()

    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    if context_text:
        combined_message = (
            f"Retrieved GDPR articles relevant to the question:\n\n{context_text}\n\n"
            f"User question: {user_text}"
        )
    else:
        combined_message = (
            "No GDPR articles were retrieved for this question -- tell the "
            f"user that plainly instead of guessing.\n\nUser question: {user_text}"
        )

    messages = (
        [{"role": "system", "content": GDPR_SYSTEM_PROMPT}]
        + clean_history
        + [{"role": "user", "content": combined_message}]
    )
    try:
        answer = _chat_backend(
            messages, model=GDPR_MODEL, num_predict=_GDPR_NUM_PREDICT,
            timeout=_GDPR_REQUEST_TIMEOUT_SECONDS,
            timeout_hint="try a narrower question about a specific article or topic.",
        )
    except RuntimeError as e:
        yield f"⚠️ {e}", gr.skip()
        return

    yield answer, _format_gdpr_sources_panel(sources)


# --- HIPAA AI (RAG over 45 CFR Parts 160 & 164) -----------------------------
#
# Same shape as GDPR AI / Canon AI just above -- pointed at HIPAA's
# regulatory text (45 CFR Part 160, General Administrative Requirements, and
# Part 164, Security and Privacy), scraped section-by-section via
# scripts/scrape_hipaa.py and embedded via scripts/embed_hipaa_to_chroma.py.
# Reuses the same generic embedding/re-ranking helpers as GDPR AI/Canon AI
# (_embed_canon_query, _rerank_candidates, _tokenize, _lexical_overlap_score)
# rather than duplicating them.
#
# One structural difference from GDPR/Canon: HIPAA's "article number"
# equivalent -- the CFR section id, e.g. "164.312" -- is a STRING, not an
# integer (unlike GDPR's plain 1..99 article_number), since CFR section
# numbers aren't sequential integers. Metadata lookups below match on that
# string directly rather than trying both str/int forms.
#
# Path/collection name here MUST match what embed_hipaa_to_chroma.py was run
# with (--chroma-dir / --collection). Adjust if you used different values.
HIPAA_CHROMA_DIR = str(Path(__file__).resolve().parent / "chroma_db")
HIPAA_COLLECTION_NAME = "hipaa"

# Same reasoning/tuning basis as _GDPR_TOP_K/_GDPR_FETCH_K -- wide enough
# that a relevant-but-not-top-1 section isn't trimmed off before the LLM
# ever sees it, without spending so much context on marginal matches.
_HIPAA_TOP_K = 6
_HIPAA_FETCH_K = 20

# Model that answers using the retrieved sections. Pinned explicitly to
# gpt-oss:20b for the same reason GDPR_MODEL/CANON_MODEL are pinned rather
# than left to inherit main.py's OLLAMA_MODEL default.
HIPAA_MODEL = "gpt-oss:20b"
_HIPAA_NUM_PREDICT = 2048
_HIPAA_REQUEST_TIMEOUT_SECONDS = 900

HIPAA_SYSTEM_PROMPT = (
    "You are a healthcare privacy / security compliance expert specializing "
    "in HIPAA (the Health Insurance Portability and Accountability Act), as "
    "implemented in its regulatory text at 45 CFR Part 160 (General "
    "Administrative Requirements) and Part 164 (Security and Privacy, "
    "covering the Privacy Rule, Security Rule, and Breach Notification "
    "Rule). Think through and answer every question strictly according to "
    "the sections provided to you below -- not from general recall of "
    "privacy/security law, and not from the law of any other jurisdiction "
    "or framework (e.g. GDPR, state privacy laws, HITECH provisions not "
    "reflected in these sections), unless the user explicitly asks about "
    "that instead.\n\n"
    "For every substantive claim, cite the specific section (and "
    "paragraph/point where relevant) you are relying on immediately after "
    "the claim -- for example: 'Covered entities must implement unique "
    "user identification for access to ePHI (45 CFR § 164.312(a)(2)(i)).' "
    "Base your answer ONLY on the sections retrieved below plus general "
    "background knowledge of HIPAA's structure -- never invent a section "
    "number or a rule that isn't in the excerpts you were given. Note "
    "where relevant whether an implementation specification is 'Required' "
    "or 'Addressable' (45 CFR § 164.306(d)) when the retrieved text says "
    "so, since that distinction materially affects compliance obligations. "
    "If the retrieved sections don't actually answer the question, say so "
    "explicitly instead of guessing -- a wrong or fabricated citation is "
    "worse than admitting the retrieval didn't cover it.\n\n"
    "Keep your answer proportionate to the question: for a broad or "
    "general topic, cover the most important, directly relevant sections "
    "rather than exhaustively enumerating every retrieved excerpt -- you "
    "can offer to go deeper on a specific section if the user wants that.\n\n"
    "You are not a substitute for advice from a qualified healthcare "
    "compliance attorney or a Privacy/Security Officer, and you should say "
    "so when a question calls for one. Always reply in the same language "
    "the user's question is written in. Never switch languages on the "
    "user unasked."
)

_HIPAA_SOURCES_PLACEHOLDER = (
    "_No sections retrieved yet. Once you ask a question, the specific "
    "HIPAA (45 CFR) sections retrieved from the vector database will be "
    "listed here._"
)
_HIPAA_SOURCES_NONE_FOUND = (
    "_Nothing was retrieved for this question -- the answer above (if any) "
    "isn't grounded in a specific section. Try rephrasing._"
)
_HIPAA_DB_MISSING_MSG = (
    "⚠️ HIPAA AI's vector database isn't available. Make sure you've run "
    "scripts/scrape_hipaa.py then scripts/embed_hipaa_to_chroma.py (pointed "
    f"at '{HIPAA_CHROMA_DIR}', collection '{HIPAA_COLLECTION_NAME}'), and "
    "that chromadb is installed (`pip install chromadb`)."
)

_hipaa_collection_cache = None


def _get_hipaa_collection():
    """Lazily connect to the persistent Chroma collection, caching the
    handle across calls. Returns None (rather than raising) on any
    failure -- callers surface a clean in-chat error instead of crashing
    the whole UI process if the vector DB hasn't been built yet."""
    global _hipaa_collection_cache
    if _hipaa_collection_cache is not None:
        return _hipaa_collection_cache
    if chromadb is None:
        return None
    try:
        client = chromadb.PersistentClient(path=HIPAA_CHROMA_DIR)
        _hipaa_collection_cache = client.get_collection(HIPAA_COLLECTION_NAME)
        return _hipaa_collection_cache
    except Exception as e:
        import traceback
        print(f"[HIPAA AI] failed to open Chroma collection at "
              f"'{HIPAA_CHROMA_DIR}' / '{HIPAA_COLLECTION_NAME}': {e}")
        traceback.print_exc()
        return None


# Matches a CFR section reference in a question: "45 CFR 164.312",
# "§ 164.312", "section 164.312", "164.312(a)(2)" -- captures just the
# leading {part}.{section} id (an optional trailing letter, e.g. "164.520a",
# is included; any paragraph suffix like "(a)(2)" is ignored for lookup
# purposes, same reasoning as GDPR's article-number regex).
_HIPAA_SECTION_NUMBER_RE = re.compile(r"\b(1(?:60|62|64)\.\d{1,4}[a-zA-Z]?)\b")


def retrieve_hipaa(query: str, n_results: int = _HIPAA_TOP_K):
    """
    Retrieves relevant HIPAA (45 CFR) sections for a query, combining two
    strategies (same overall approach as retrieve_gdpr/retrieve_canons):

    1. Exact match: if the query names a specific section id (e.g.
       "164.312" or "§ 160.103"), that section is fetched directly via a
       metadata filter and placed first.
    2. Semantic search + hybrid re-rank: the query is embedded with
       nomic-embed-text (via the shared _embed_canon_query helper),
       _HIPAA_FETCH_K candidates are pulled by vector similarity, then
       re-scored by blending vector similarity with lexical keyword
       overlap and filtered for diversity (via the shared
       _rerank_candidates helper), before trimming down to n_results.

    Returns (context_text, sources) -- or (None, None) on any failure
    (missing DB, unreachable Ollama, etc.), which callers turn into a
    clear in-chat message rather than a silent empty answer.
    """
    collection = _get_hipaa_collection()
    if collection is None:
        return None, None

    exact_docs, exact_metas = [], []
    for section_id in dict.fromkeys(_HIPAA_SECTION_NUMBER_RE.findall(query)):  # dedupe, keep order
        try:
            hit = collection.get(where={"section_id": section_id})
        except Exception:
            continue
        for doc, meta in zip(hit.get("documents") or [], hit.get("metadatas") or []):
            exact_docs.append(doc)
            exact_metas.append(meta)

    query_vector = None
    if _has_meaningful_content(query):
        query_vector = _embed_canon_query(query)  # generic nomic-embed-text helper, reused as-is
        if query_vector is None and not exact_docs:
            return None, None
    elif not exact_docs:
        return "", []  # greeting/chit-chat, no real topic to search on and no exact citation either

    vector_docs, vector_metas = [], []
    if query_vector is not None:
        try:
            results = collection.query(
                query_embeddings=[query_vector],
                n_results=_HIPAA_FETCH_K,
                include=["documents", "metadatas", "distances", "embeddings"],
            )
            raw_docs = (results.get("documents") or [[]])[0]
            raw_metas = (results.get("metadatas") or [[]])[0]
            raw_dists = (results.get("distances") or [[]])[0]
            raw_embeds = (_embeddings_to_list(results.get("embeddings")) or [[]])[0]
            vector_docs, vector_metas = _rerank_candidates(
                query, query_vector, raw_docs, raw_metas, raw_embeds,
                top_n=n_results, min_similarity=_MIN_RELEVANCE_SIMILARITY,
                label="HIPAA AI",
            )
        except Exception:
            pass  # exact matches (if any) still make it through below

    # Merge exact matches first, then vector results, de-duping by section
    # id so a section named explicitly doesn't also show up twice.
    seen = set()
    documents, metadatas = [], []
    for doc, meta in zip(exact_docs + vector_docs, exact_metas + vector_metas):
        key = meta.get("section_id", "?")
        if key in seen:
            continue
        seen.add(key)
        documents.append(doc)
        metadatas.append(meta)

    if not documents:
        return "", []

    context_parts = []
    sources = []
    for doc, meta in zip(documents, metadatas):
        section_id = meta.get("section_id", "?")
        title = meta.get("title", "")
        path = meta.get("hierarchy_path", "")
        url = meta.get("source_url", "")
        cross_refs = meta.get("cross_references", "")
        context_parts.append(f"[45 CFR § {section_id} -- {title}] ({path})\n{doc}")
        sources.append({
            "section_id": section_id, "title": title, "hierarchy_path": path,
            "source_url": url, "cross_references": cross_refs,
        })

    context_text = "\n\n---\n\n".join(context_parts)
    return context_text, sources


def _format_hipaa_sources_panel(sources: list[dict] | None) -> str:
    if not sources:
        return _HIPAA_SOURCES_NONE_FOUND
    lines = []
    for s in sources:
        line = f"- **45 CFR § {s['section_id']}** — {s['title']}"
        if s.get("hierarchy_path"):
            line += f"  \n  {s['hierarchy_path']}"
        if s.get("source_url"):
            line += f"  \n  [{s['source_url']}]({s['source_url']})"
        if s.get("cross_references"):
            line += f"  \n  🔗 Cross-references: {s['cross_references']}"
        lines.append(line)
    return "\n\n".join(lines)


def hipaa_chat_fn(message, history):
    """
    RAG chat over the scraped/embedded HIPAA regulatory text. Not
    multimodal (no file attachment) -- same reasoning as
    canon_chat_fn/gdpr_chat_fn: the point of this tab is grounding answers
    in the pre-built vector database, not whatever the user happens to
    attach.

    Returns (answer, sources_panel_markdown) -- see hipaa_sources_panel's
    wiring below, same additional_outputs pattern as Canon AI / GDPR AI.
    """
    user_text = message.get("text", message) if isinstance(message, dict) else message

    if chromadb is None or _get_hipaa_collection() is None:
        yield _HIPAA_DB_MISSING_MSG, gr.skip()
        return

    yield "🔍 *Searching ChromaDB for HIPAA sections relevant to your question…*", gr.skip()

    context_text, sources = retrieve_hipaa(user_text)
    if context_text is None:
        yield _HIPAA_DB_MISSING_MSG, gr.skip()
        return

    n_found = len(sources) if sources else 0
    if n_found:
        yield (
            f"🏥 *Found {n_found} relevant section(s) — asking the model to "
            f"answer using them…*",
            gr.skip(),
        )
    else:
        yield "🤔 *Nothing matched in the vector database — asking the model how to proceed…*", gr.skip()

    clean_history = [
        {"role": turn["role"], "content": turn["content"]}
        for turn in history
        if isinstance(turn.get("content"), str)
    ]

    if context_text:
        combined_message = (
            f"Retrieved HIPAA (45 CFR) sections relevant to the question:\n\n{context_text}\n\n"
            f"User question: {user_text}"
        )
    else:
        combined_message = (
            "No HIPAA sections were retrieved for this question -- tell "
            f"the user that plainly instead of guessing.\n\nUser question: {user_text}"
        )

    messages = (
        [{"role": "system", "content": HIPAA_SYSTEM_PROMPT}]
        + clean_history
        + [{"role": "user", "content": combined_message}]
    )
    try:
        answer = _chat_backend(
            messages, model=HIPAA_MODEL, num_predict=_HIPAA_NUM_PREDICT,
            timeout=_HIPAA_REQUEST_TIMEOUT_SECONDS,
            timeout_hint="try a narrower question about a specific section or topic.",
        )
    except RuntimeError as e:
        yield f"⚠️ {e}", gr.skip()
        return

    yield answer, _format_hipaa_sources_panel(sources)


# --- Knesset law RAG (grounds the Attorney/DictaLM tab) -------------------
#
# Same shape as Canon AI/GDPR AI/HIPAA AI above, pointed at Israeli
# primary legislation pulled from the Knesset's official OData API
# (scripts/scrape_knesset_laws.py) and embedded by
# scripts/embed_knesset_to_chroma.py. Reuses the generic
# _embed_canon_query/_rerank_candidates/_tokenize helpers already defined
# above rather than duplicating them.
#
# UNLIKE Canon/GDPR/HIPAA, this collection has NO exact-number lookup.
# Canon numbers and GDPR/HIPAA article/section numbers are unique within
# their single corpus; Israeli "סעיף 1" exists in hundreds of different
# laws, so a bare section-number match can't disambiguate which law is
# meant. Retrieval here is semantic+lexical hybrid rerank only -- see
# scripts/embed_knesset_to_chroma.py's module docstring for the full
# reasoning.
#
# Path/collection name MUST match what embed_knesset_to_chroma.py was run
# with.
KNESSET_CHROMA_DIR = str(Path(__file__).resolve().parent / "chroma_db")
KNESSET_COLLECTION_NAME = "knesset_laws"

_KNESSET_TOP_K = 6
_KNESSET_FETCH_K = 20

# --- Whole-law retrieval ---------------------------------------------------
#
# The plain top-K semantic search above has a real blind spot: _KNESSET_TOP_K
# is a GLOBAL cap across the entire collection, not per-law. A law with more
# sections than that has no guarantee its most relevant sections all make
# the cut -- they're competing against every other embedded law's chunks for
# the same 6 slots, and a section that's relevant to the law overall but
# phrased differently than the question can simply lose. For a question like
# "summarize this whole law" or "what does it say about X, Y, and Z" (where
# the real answer spans many sections), that silently produces a confident
# but incomplete answer with no indication anything was left out.
#
# Fix: before falling back to semantic search, check whether the question
# clearly NAMES one specific law (by real title-word overlap, not vector
# similarity -- vector similarity is exactly the mechanism with the blind
# spot above, so it can't be what decides whether to bypass it). If so,
# fetch EVERY chunk for that law_id directly via a metadata filter, sorted
# by section number, instead of competing for top-K slots at all.
#
# This is a heuristic, not a guarantee -- a question that doesn't clearly
# name the law falls through to the existing semantic search unchanged.

# Fraction of a law's title's meaningful words that must appear in the
# question before that law is treated as "clearly named" rather than just
# topically related. Deliberately high: a false positive here means
# potentially dumping thousands of characters of the WRONG law into
# DictaLM's context, which is worse than just falling through to the
# (already reasonable) semantic search.
_KNESSET_WHOLE_LAW_TOKEN_OVERLAP_THRESHOLD = 0.7
# A title with fewer meaningful words than this can't be matched
# confidently at all (e.g. a 1-2-word title could trivially "match" lots
# of unrelated questions) -- such laws simply never qualify for whole-law
# mode and always go through semantic search instead.
_KNESSET_WHOLE_LAW_MIN_TITLE_TOKENS = 3

# Budget for a whole-law fetch, in WEIGHTED units (see _estimate_token_weight
# below), not raw characters. Previously this was a flat 20000-CHARACTER cap
# assuming ~3 chars/token -- that assumption was never actually verified
# against DictaLM's real tokenizer behavior on Hebrew text, and if it's
# wrong in the optimistic direction, the real prompt can exceed
# _LEGAL_NUM_CTX (16384) once system prompt + history + output budget
# (_LEGAL_NUM_PREDICT, 6144) are added on top. That's not a soft failure --
# see main.py's _num_ctx_for comment: when a real ceiling isn't reachable,
# Ollama falls back to context-shifting instead of stopping cleanly, and
# generation can run for a very long time instead of erroring, which looks
# exactly like "takes ages and doesn't answer."
#
# Sized conservatively on purpose: available input budget is roughly
# _LEGAL_NUM_CTX (11264, corrected -- see that constant's comment above for
# why it was 16384 before and that was the actual root cause of the
# original "never answers" symptom) minus _LEGAL_NUM_PREDICT's output
# reserve (6144) minus headroom for system prompt/history/question
# (~1500 tokens) = ~3600 tokens. 3200 weighted units stays under that even
# under a pessimistic per-token weight -- see _estimate_token_weight's
# docstring for why Hebrew content specifically can't be trusted to a flat
# chars-per-token guess. If you confirm empirically (via Ollama's own
# server log, which prints actual prompt token counts per request) that
# your setup has more headroom than this, it's safe to raise -- but raise
# it in small steps and re-test, since this exact miscalibration (budget
# sized against a context window the code wasn't actually requesting) is
# what caused the original hang.
_KNESSET_WHOLE_LAW_MAX_WEIGHT = 3200

# --- Chunked (map-reduce) whole-law answering ------------------------------
#
# _fetch_whole_law below still exists (kept for reference / in case it's
# useful elsewhere) but is no longer called by the Attorney tabs. The
# primary whole-law path now goes through _fetch_whole_law_sections +
# _answer_whole_law_chunked instead of stuffing the entire law into one
# oversized prompt. Confirmed directly against the Ollama server log (Aug
# 10): a single 5180-token whole-law prompt took 113s of CPU prefill and
# came back with only 64 tokens of output -- a small model doesn't reason
# well over one huge undifferentiated context, and the old weight-based
# truncation also meant a law longer than budget silently lost its later
# sections. Processing the law in small batches (map) then synthesizing a
# final answer from only the relevant extracts (reduce) fixes both: no
# length ceiling, and the model only ever has to reason over a small,
# focused piece of text at a time.
_MAP_BATCH_MAX_WEIGHT = 3000        # weighted units per map-step batch (see _estimate_token_weight)
_MAP_NUM_PREDICT = 500              # map steps only extract, they don't need to reason at length
_MAP_NUM_CTX = 6144                 # covers a batch's real tokens + system prompt + output, comfortably
_MAP_REQUEST_TIMEOUT_SECONDS = 600  # each batch is small; one batch taking 10min means something's wrong

_MAP_SYSTEM_PROMPT = (
    "You are extracting relevant material from a section of an Israeli law "
    "to help answer a legal question. You will be given some sections "
    "(סעיפים) of the law and a question. If any of these specific sections "
    "are relevant, quote or closely paraphrase the relevant part and note "
    "the סעיף number. If NONE of these sections are relevant, respond with "
    "EXACTLY: לא רלוונטי\n\n"
    "Be concise -- this is one batch of many. Do not try to answer the "
    "full question yet; only report what THIS batch contains."
)
_MAP_IRRELEVANT_MARKER = "לא רלוונטי"

# Matches Hebrew, English, or numeric tokens for lightweight lexical
# matching between a question and a law's title -- NOT the same as
# _tokenize() above, which is Latin-script-only (built for Canon AI's
# Italian text) and would silently match nothing on Hebrew titles.
_KNESSET_TITLE_TOKEN_RE = re.compile(r"[א-ת]+|[a-zA-Z]+|\d+")


def _knesset_title_tokens(text: str) -> set[str]:
    return {t for t in _KNESSET_TITLE_TOKEN_RE.findall(text or "") if len(t) > 1}


def _estimate_token_weight(text: str) -> float:
    """Rough, script-aware proxy for prompt-token cost, deliberately
    pessimistic. General-purpose BPE tokenizers are usually trained on
    corpora dominated by Latin-script text, and multi-byte-UTF-8 scripts
    like Hebrew routinely need MORE tokens per character than English,
    not fewer -- the previous flat "~3 chars/token" guess this budget used
    had no actual verification behind it for this specific model/script
    combination. Weighting Hebrew characters ~2x more heavily than other
    characters is still a guess, but an intentionally conservative one:
    better to under-fill the context window (a smaller-than-necessary
    excerpt, clearly flagged as truncated) than to overflow it (a
    request that appears to hang -- see _KNESSET_WHOLE_LAW_MAX_WEIGHT's
    comment above for what that failure mode actually looks like)."""
    hebrew_chars = sum(1 for ch in text if 0x0590 <= ord(ch) <= 0x05FF)
    other_chars = len(text) - hebrew_chars
    return hebrew_chars * 2.0 + other_chars * 1.0


_knesset_law_titles_cache = {"count": None, "map": {}}


def _get_knesset_law_titles(collection) -> dict:
    """Returns {law_id: title} for every law currently embedded. Cached and
    invalidated by collection.count() changing (e.g. after re-running
    embed_knesset_to_chroma.py / embed_local_law_pdfs.py) rather than
    re-fetched on every single question -- this is O(all chunks) since
    Chroma has no native "distinct" query, which is fine at this app's
    scale (a personal/small-team legal database) but would need revisiting
    at a much larger corpus size."""
    global _knesset_law_titles_cache
    try:
        current_count = collection.count()
    except Exception:
        current_count = None
    if current_count is not None and _knesset_law_titles_cache["count"] == current_count:
        return _knesset_law_titles_cache["map"]
    try:
        all_rows = collection.get(include=["metadatas"])
    except Exception as e:
        print(f"[Knesset RAG] couldn't refresh law-title map: {e}")
        return _knesset_law_titles_cache["map"]  # stale is better than nothing
    law_map = {}
    for meta in all_rows.get("metadatas") or []:
        law_id = meta.get("law_id")
        title = meta.get("title")
        if law_id and title:
            law_map[law_id] = title
    _knesset_law_titles_cache = {"count": current_count, "map": law_map}
    return law_map


def _detect_named_law(query: str, law_map: dict) -> tuple[str, str, float] | None:
    """Returns (law_id, title, overlap_score) for the best-matching law if
    the question clearly names it, else None. See the module comment above
    _KNESSET_WHOLE_LAW_TOKEN_OVERLAP_THRESHOLD for why this is intentionally
    conservative (a false positive is worse than a false negative here)."""
    query_tokens = _knesset_title_tokens(query)
    if not query_tokens:
        return None
    best = None
    for law_id, title in law_map.items():
        title_tokens = _knesset_title_tokens(title)
        if len(title_tokens) < _KNESSET_WHOLE_LAW_MIN_TITLE_TOKENS:
            continue
        overlap = len(title_tokens & query_tokens) / len(title_tokens)
        if overlap < _KNESSET_WHOLE_LAW_TOKEN_OVERLAP_THRESHOLD:
            continue
        if best is None or overlap > best[2]:
            best = (law_id, title, overlap)
    return best


def _section_sort_key(meta: dict):
    """Sorts a law's chunks into a sensible reading order: preamble/no
    section-number chunks first, then numeric section order (not string
    order, which would put "10" before "2"). A trailing Hebrew letter on a
    section number (e.g. "12א", an inserted amendment section) sorts right
    after its base number via the secondary key."""
    section = (meta.get("section_number") or "").strip()
    if not section:
        return (0, 0, "")
    m = re.match(r"(\d+)([א-ת]?)", section)
    if not m:
        return (1, 0, section)
    return (1, int(m.group(1)), m.group(2))


def _fetch_whole_law(collection, law_id: str, title: str, max_weight: float = _KNESSET_WHOLE_LAW_MAX_WEIGHT):
    """Fetches EVERY chunk for one law_id directly (bypassing top-K
    semantic search entirely), sorted into section order, truncated by
    weighted budget (never mid-section, see _estimate_token_weight) if the
    law is longer than max_weight. Returns (context_text, sources,
    truncated) -- context_text is "" (not None) if the law_id somehow has
    no chunks, so callers can distinguish "found the law but it's empty"
    from "the fetch itself failed".

    max_weight defaults to the 24B tab's budget but is passed explicitly by
    retrieve_knesset_laws for callers with a different context window (see
    the Attorney 1.7B tab's _KNESSET_FAST_WHOLE_LAW_MAX_WEIGHT) -- each
    tab's num_ctx is different, so a single hardcoded budget can't be
    correct for both."""
    try:
        result = collection.get(where={"law_id": law_id}, include=["documents", "metadatas"])
    except Exception as e:
        print(f"[Knesset RAG] whole-law fetch failed for law_id={law_id}: {e}")
        return None, None, False

    docs = result.get("documents") or []
    metas = result.get("metadatas") or []
    if not docs:
        return "", [], False

    paired = sorted(zip(docs, metas), key=lambda dm: _section_sort_key(dm[1]))

    included, total_weight, total_chars, truncated = [], 0.0, 0, False
    for doc, meta in paired:
        weight = _estimate_token_weight(doc)
        if included and total_weight + weight > max_weight:
            truncated = True
            break
        included.append((doc, meta))
        total_weight += weight
        total_chars += len(doc)

    context_parts, sources = [], []
    for doc, meta in included:
        section = meta.get("section_number", "")
        label = f"[{title}" + (f", סעיף {section}" if section else "") + "]"
        context_parts.append(f"{label}\n{doc}")
        sources.append({"title": title, "section_number": section, "source_url": meta.get("source_url", "")})

    if truncated:
        context_parts.append(
            f"[NOTE: this law has more sections than fit in this answer -- only the "
            f"first {len(included)} section(s) (in section-number order) are included "
            f"above. If the user's question concerns a later section, say so explicitly "
            f"rather than answering as if the whole law was reviewed.]"
        )

    # Printed unconditionally (not just on truncation) so a slow/stuck
    # question is actually diagnosable from the terminal instead of a
    # guess: if you see this line print a large weight/char count right
    # before a long hang, that's a strong signal the prompt size (not a
    # ChromaDB lock or something else entirely) is the bottleneck --
    # compare against Ollama's own server log, which prints the ACTUAL
    # prompt token count it received, to check whether this budget's
    # weighting estimate is too conservative or not conservative enough.
    print(f"[Knesset RAG] whole-law fetch for {title!r}: {len(included)}/{len(paired)} "
          f"section(s), {total_chars} chars, ~{total_weight:.0f} weighted units "
          f"(budget: {_KNESSET_WHOLE_LAW_MAX_WEIGHT}){' -- TRUNCATED' if truncated else ''}")

    return "\n\n---\n\n".join(context_parts), sources, truncated


def _fetch_whole_law_sections(collection, law_id: str, title: str):
    """Fetches EVERY chunk for one law_id, sorted into section order, with
    NO weight-based truncation -- feeds _answer_whole_law_chunked, which
    processes arbitrarily long laws across multiple bounded calls instead
    of cutting the law off the way _fetch_whole_law's single-prompt budget
    did. Returns a list of (doc, meta) pairs, or None if the fetch itself
    failed (distinct from an empty list, which means the law_id genuinely
    has no chunks)."""
    try:
        result = collection.get(where={"law_id": law_id}, include=["documents", "metadatas"])
    except Exception as e:
        print(f"[Knesset RAG] whole-law section fetch failed for law_id={law_id}: {e}")
        return None
    docs = result.get("documents") or []
    metas = result.get("metadatas") or []
    if not docs:
        return []
    return sorted(zip(docs, metas), key=lambda dm: _section_sort_key(dm[1]))


def _batch_sections(sections: list[tuple[str, dict]], max_weight: float) -> list[list[tuple[str, dict]]]:
    """Groups (doc, meta) pairs into batches under max_weight (see
    _estimate_token_weight), never splitting a single section across two
    batches -- same 'never truncate mid-section' principle _fetch_whole_law
    already used, just applied per-batch instead of per-whole-law-budget."""
    batches, current, current_weight = [], [], 0.0
    for doc, meta in sections:
        weight = _estimate_token_weight(doc)
        if current and current_weight + weight > max_weight:
            batches.append(current)
            current, current_weight = [], 0.0
        current.append((doc, meta))
        current_weight += weight
    if current:
        batches.append(current)
    return batches


def _answer_whole_law_chunked(
    user_text: str, title: str, sections: list[tuple[str, dict]],
    clean_history: list[dict], file_context: str,
    *, model: str, num_predict: int, num_ctx: int, request_timeout: int,
    reduce_max_weight: float,
):
    """
    Answers a question grounded in an ENTIRE law via map-reduce instead of
    one oversized prompt (see the module comment above _MAP_BATCH_MAX_WEIGHT
    for the full reasoning):

      map:    the law is split into small batches; each batch gets its own
              short, narrow call asking only "is anything here relevant,
              and if so what" -- cheap, bounded, and small enough that the
              model can actually attend to it properly instead of skimming
              a single multi-thousand-token blob.
      reduce: only the batches that came back relevant (already condensed
              by the map step) are combined into ONE final call using the
              normal LEGAL_SYSTEM_PROMPT, producing the real answer with
              citations -- same shape as the existing non-chunked path.

    This is a generator: yields (status, gr.skip(), gr.skip()) progress
    tuples matching _legal_chat_fn_impl's existing shape while it works
    through the batches, then RETURNS (answer, sources, retrieval_info) via
    the generator-return protocol -- callers use:
        result = yield from _answer_whole_law_chunked(...)
    An empty-string answer in the returned tuple means every batch came
    back irrelevant (a real possibility: the law was found by name, but
    the specific question isn't actually answered by anything in it) --
    callers should fall through to semantic search rather than treating
    that as a final answer.
    """
    batches = _batch_sections(sections, _MAP_BATCH_MAX_WEIGHT)
    n_batches = len(batches)
    relevant_extracts, sources = [], []

    for i, batch in enumerate(batches):
        yield (f"⚖️ *Reading {title} — section batch {i + 1}/{n_batches}…*", gr.skip(), gr.skip())

        batch_text = "\n\n---\n\n".join(
            f"[{title}" + (f", סעיף {meta.get('section_number')}" if meta.get("section_number") else "") + f"]\n{doc}"
            for doc, meta in batch
        )
        map_messages = [
            {"role": "system", "content": _MAP_SYSTEM_PROMPT},
            {"role": "user", "content": f"Question: {user_text}\n\nSections:\n\n{batch_text}"},
        ]
        try:
            extract = _chat_backend(
                map_messages, model=model, num_predict=_MAP_NUM_PREDICT,
                num_ctx=_MAP_NUM_CTX, timeout=_MAP_REQUEST_TIMEOUT_SECONDS,
                timeout_hint="try a narrower question.",
            )
        except RuntimeError as e:
            print(f"[Knesset RAG map-reduce] {title!r} batch {i + 1}/{n_batches} failed: {e}")
            continue  # one bad batch shouldn't sink the whole answer

        if _MAP_IRRELEVANT_MARKER in extract and len(extract.strip()) < 40:
            continue

        relevant_extracts.append(extract.strip())
        for _doc, meta in batch:
            sources.append({
                "title": title,
                "section_number": meta.get("section_number", ""),
                "source_url": meta.get("source_url", ""),
            })

    if not relevant_extracts:
        print(f"[Knesset RAG map-reduce] {title!r}: none of {n_batches} batch(es) "
              "came back relevant to the question.")
        return "", [], {"mode": "whole_law", "truncated": False, "title": title}

    # Reduce, with its own (much more forgiving) safety cap -- map output is
    # already condensed, so this rarely triggers, but it's a real net.
    combined, total_weight = [], 0.0
    for extract in relevant_extracts:
        weight = _estimate_token_weight(extract)
        if combined and total_weight + weight > reduce_max_weight:
            break
        combined.append(extract)
        total_weight += weight

    law_context = "\n\n---\n\n".join(combined)
    yield (f"⚖️ *Combining {len(combined)} relevant excerpt(s) from {title} — consulting AI…*",
           gr.skip(), gr.skip())

    message_parts = [
        f"Retrieved and pre-filtered excerpts from the full text of {title}, "
        f"relevant to the question below:\n\n{law_context}"
    ]
    if file_context:
        message_parts.append(f"The user attached the following document(s):\n\n{file_context}")
    message_parts.append(f"User question: {user_text}")

    messages = (
        [{"role": "system", "content": LEGAL_SYSTEM_PROMPT}]
        + clean_history
        + [{"role": "user", "content": "\n\n".join(message_parts)}]
    )
    answer = _chat_backend(
        messages, model=model, num_predict=num_predict, num_ctx=num_ctx,
        timeout=request_timeout,
        timeout_hint="try a narrower question -- e.g. ask about a specific section rather than a whole law.",
    )
    return answer, sources, {
        "mode": "whole_law", "truncated": False, "title": title,
        "chunked": True, "n_batches": n_batches,
    }


_knesset_collection_cache = None


def _get_knesset_collection():
    """Same lazy-connect-and-cache pattern as _get_canon_collection/
    _get_gdpr_collection/_get_hipaa_collection -- returns None (not an
    exception) if the DB hasn't been built yet, so the Attorney tab
    degrades to DictaLM's own knowledge instead of breaking."""
    global _knesset_collection_cache
    if _knesset_collection_cache is not None:
        return _knesset_collection_cache
    if chromadb is None:
        return None
    try:
        client = chromadb.PersistentClient(path=KNESSET_CHROMA_DIR)
        _knesset_collection_cache = client.get_collection(KNESSET_COLLECTION_NAME)
        return _knesset_collection_cache
    except Exception as e:
        import traceback
        print(f"[Knesset RAG] failed to open Chroma collection at "
              f"'{KNESSET_CHROMA_DIR}' / '{KNESSET_COLLECTION_NAME}': {e}")
        traceback.print_exc()
        return None


def retrieve_knesset_laws(
    query: str,
    n_results: int = _KNESSET_TOP_K,
    whole_law_max_weight: float = _KNESSET_WHOLE_LAW_MAX_WEIGHT,
):
    """
    Retrieves relevant Israeli statute excerpts for a query via semantic
    (hybrid vector+lexical) search only.

    NOTE: whole_law_max_weight is kept as a parameter for backward
    compatibility with existing callers, but this function no longer has
    its own whole-law branch -- that logic now lives in _legal_chat_fn_impl
    (via _detect_named_law + _fetch_whole_law_sections +
    _answer_whole_law_chunked), which processes a named law's FULL text
    through map-reduce instead of one truncated prompt. Doing whole-law
    detection here too would risk re-triggering the old truncating
    single-prompt fetch as a "fallback" right after the chunked path
    already determined nothing in the law was relevant -- keeping this
    function semantic-search-only avoids that double-fetch/conflicting-
    logic trap. _fetch_whole_law itself is left in place (unused by this
    function, but harmless) in case it's useful for something else later.

    Returns (context_text, sources, retrieval_info) where retrieval_info is
    a dict with "mode" ("semantic" | None) and "truncated" (bool) -- or
    (None, None, {"reason": ...}) on failure, where reason is
    "chroma_unavailable" (the vector DB itself couldn't be opened -- see
    _get_knesset_collection's own log line for the traceback) or
    "embedding_unavailable" (ChromaDB is fine, but the call to Ollama for
    the query embedding failed/timed out -- see _embed_canon_query's log
    line). Distinguishing these used to be impossible from the caller's
    side: both collapsed into one identical "database unavailable" message
    with no way to tell a DB problem from an Ollama problem without reading
    server-side logs. Callers turn either failure into "answer from general
    knowledge, no grounding available" rather than a hard error -- unlike
    the Canon/GDPR/HIPAA tabs, the Attorney tab is expected to still work
    (just less precisely) when this DB isn't present.
    """
    collection = _get_knesset_collection()
    if collection is None:
        return None, None, {"reason": "chroma_unavailable"}

    # Same content-free-query guard as retrieve_canons/retrieve_gdpr/
    # retrieve_hipaa (see _has_meaningful_content's docstring), but using
    # _knesset_title_tokens instead of _tokenize -- this corpus is Hebrew,
    # and _tokenize's [a-z'] regex is Latin-only (built for Canon AI's
    # Italian text), so it would silently treat every Hebrew query as
    # content-free. A short length-3+ requirement (stricter than
    # _knesset_title_tokens' own >1 filter, which alone would let "hi"
    # through) keeps this catching plain greetings without a Hebrew
    # stopword list.
    if not any(len(t) > 2 for t in _knesset_title_tokens(query)):
        return "", [], {"mode": "semantic", "truncated": False}

    query_vector = _embed_canon_query(query)  # generic nomic-embed-text helper, reused as-is
    if query_vector is None:
        return None, None, {"reason": "embedding_unavailable"}

    try:
        results = collection.query(
            query_embeddings=[query_vector],
            n_results=_KNESSET_FETCH_K,
            include=["documents", "metadatas", "distances", "embeddings"],
        )
        raw_docs = (results.get("documents") or [[]])[0]
        raw_metas = (results.get("metadatas") or [[]])[0]
        raw_dists = (results.get("distances") or [[]])[0]
        raw_embeds = (_embeddings_to_list(results.get("embeddings")) or [[]])[0]
        documents, metadatas = _rerank_candidates(
            query, query_vector, raw_docs, raw_metas, raw_embeds,
            top_n=n_results, min_similarity=_MIN_RELEVANCE_SIMILARITY,
            label="Knesset RAG",
        )
    except Exception as e:
        print(f"[Knesset RAG] query failed: {e}")
        return None, None, {}

    if not documents:
        return "", [], {"mode": "semantic", "truncated": False}

    context_parts = []
    sources = []
    for doc, meta in zip(documents, metadatas):
        title = meta.get("title", "")
        section = meta.get("section_number", "")
        url = meta.get("source_url", "")
        label = f"[{title}" + (f", סעיף {section}" if section else "") + "]"
        context_parts.append(f"{label}\n{doc}")
        sources.append({"title": title, "section_number": section, "source_url": url})

    context_text = "\n\n---\n\n".join(context_parts)
    return context_text, sources, {"mode": "semantic", "truncated": False}


def _format_knesset_sources_panel(sources: list[dict] | None, retrieval_info: dict | None = None) -> str:
    if not sources:
        return _KNESSET_SOURCES_NONE_FOUND
    header = ""
    if retrieval_info and retrieval_info.get("mode") == "whole_law":
        header = f"📖 **Full text retrieved: {retrieval_info.get('title', '')}** ({len(sources)} section(s))"
        if retrieval_info.get("truncated"):
            header += (
                "\n\n⚠️ *This law is longer than fits in one answer -- only the sections "
                "above (by section number) were included. Ask about a later section "
                "specifically to see the rest.*"
            )
        header += "\n\n---\n\n"
    lines = []
    for s in sources:
        line = f"- **{s['title']}**"
        if s.get("section_number"):
            line += f" — סעיף {s['section_number']}"
        if s.get("source_url"):
            line += f"  \n  [{s['source_url']}]({s['source_url']})"
        lines.append(line)
    return header + "\n\n".join(lines)


_KNESSET_SOURCES_PLACEHOLDER = (
    "_No statutes retrieved yet. Once you ask a question, any Israeli "
    "law excerpts retrieved from the Knesset legislation database will "
    "be listed here._"
)
_KNESSET_SOURCES_NONE_FOUND = (
    "_Nothing matched in the Knesset law database for this question -- "
    "the answer relies on the Agent's own knowledge only, not a retrieved "
    "statute. Verify independently._"
)

# Appended to the end of EVERY Attorney-tab answer (not just shown once
# above the chat like the tab's gr.Markdown description) -- so the
# not-a-lawyer + Knesset-sourcing notice survives a copy-paste or
# screenshot of just the answer text. Same placement pattern as the
# existing _NO_CITATION_NOTE, and deliberately says something slightly
# different depending on whether anything was actually retrieved for this
# specific answer, rather than one generic disclaimer for every case.
_KNESSET_DISCLAIMER_GROUNDED = (
    "\n\n---\n⚖️ *This answer draws on statute excerpts retrieved from the "
    "Knesset's official legislation database (see \"Statutes retrieved\" "
    "in the sidebar). It is not legal advice and is not a substitute for "
    "a licensed attorney -- verify anything consequential against the "
    "actual legislation.*"
)
_KNESSET_DISCLAIMER_UNGROUNDED = (
    "\n\n---\n⚖️ *No matching statute was retrieved from the Knesset "
    "database for this question -- this answer relies on the model's own "
    "general knowledge, not a verified source. It is not legal advice and "
    "is not a substitute for a licensed attorney.*"
)
# Fires only when BOTH danger signs are present at once -- see the
# combined-signal check in legal_chat_fn above this note is used from.
# Deliberately more alarmed in tone (🚨 vs ⚖️) than the two disclaimers
# above: this is the specific case where the model produced something
# that reads as an authoritative citation with nothing behind it, which is
# a materially worse failure mode than an answer that's honestly vague.
_KNESSET_DISCLAIMER_UNGROUNDED_WITH_CITATIONS = (
    "\n\n---\n🚨 *No matching statute was found in the Knesset database for this "
    "question, yet this answer includes what looks like a specific legal citation "
    "(see \"Citations found\" in the sidebar). Small local models can and do "
    "fabricate plausible-looking section numbers and case names when they aren't "
    "grounded in retrieved text -- treat any citation in this answer as UNVERIFIED "
    "until you check it against the actual legislation. This is not legal advice "
    "and is not a substitute for a licensed attorney.*"
)


def set_hebrew_from_source_lang(source_lang):
    """
    Fires on the Translate tab's Source language dropdown. Hebrew/Tesseract
    routing is now fully automatic — there's no checkbox for the user to
    manage. Returns (is_hebrew, status_markdown_update).

    NOTE: assumes the LANGUAGES dict in app/translate.py uses the key
    "hebrew" (matching the lowercase convention "english"/"spanish" already
    used as defaults below) — double check that key against translate.py if
    this doesn't trigger correctly.
    """
    is_hebrew = (source_lang or "").strip().lower() == "hebrew"
    if is_hebrew:
        return True, gr.update(
            value="🔤 **Hebrew selected — Tesseract OCR will be used automatically** "
                  "(the only engine here with Hebrew support).",
            visible=True,
        )
    return False, gr.update(value="", visible=False)


LOGO_PATH = "app/assets/logo_t.png"  # transparent logo — put your logo file here, any size, it's auto-resized below

SUPPORTED_INVOICE_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}


def split_pdf_to_pages(pdf_path, output_dir):
    """Splits a multi-page PDF into one single-page PDF per page, so each
    page is later OCR'd/classified as its own separate document."""
    reader = PdfReader(str(pdf_path))
    base_name = Path(pdf_path).stem
    page_paths = []
    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        out_path = Path(output_dir) / f"{base_name}_page{i + 1}.pdf"
        with open(out_path, "wb") as f:
            writer.write(f)
        page_paths.append(out_path)
    return page_paths


def expand_pdfs_to_pages(files, work_dir):
    """Given a mixed list of files, splits every PDF into per-page PDFs
    and leaves image files untouched (they're already single-page)."""
    expanded = []
    for f in files:
        if f.suffix.lower() == ".pdf":
            expanded.extend(split_pdf_to_pages(f, work_dir))
        else:
            expanded.append(f)
    return expanded


def process_invoices(uploaded_file, company_name, hebrew_batch, progress=gr.Progress()):
    if uploaded_file is None:
        return None, "Please upload a ZIP or PDF file first."

    work_dir = tempfile.mkdtemp()
    suffix = Path(uploaded_file.name).suffix.lower()

    if suffix == ".zip":
        try:
            with zipfile.ZipFile(uploaded_file.name, "r") as zf:
                zf.extractall(work_dir)
        except zipfile.BadZipFile:
            return None, "That file doesn't look like a valid ZIP archive."
        raw_files = sorted(
            p for p in Path(work_dir).rglob("*")
            if p.is_file() and p.suffix.lower() in SUPPORTED_INVOICE_EXTS
        )
        files = expand_pdfs_to_pages(raw_files, work_dir)
    elif suffix == ".pdf":
        files = split_pdf_to_pages(Path(uploaded_file.name), work_dir)
    else:
        return None, f"Unsupported file type '{suffix}'. Upload a .zip or a .pdf."

    if not files:
        return None, "No supported files/pages found."

    sales_rows, expense_rows, unrecognized = [], [], []

    for i, fpath in enumerate(files):
        progress((i + 1) / len(files), desc=f"Processing {fpath.name} ({i + 1}/{len(files)})")
        try:
            with open(fpath, "rb") as f:
                extract_resp = requests.post(
                    f"{BACKEND_URL}/extract-text", files={"file": f}, data={"hebrew": hebrew_batch}
                )
            extract_resp.raise_for_status()
            markdown_text = extract_resp.json()["markdown"]

            classify_resp = requests.post(
                f"{BACKEND_URL}/classify-invoice",
                json={"markdown": markdown_text, "filename": fpath.name, "company_name": company_name},
            )
            classify_resp.raise_for_status()
            result = classify_resp.json()
        except Exception:
            unrecognized.append(fpath.name)
            continue

        doc_type = result.get("document_type", "unrecognized")
        row = [
            result.get("filename", fpath.name),
            result.get("date", ""),
            result.get("party_name", ""),
            result.get("invoice_number", ""),
            result.get("amount", 0),
            result.get("vat", 0),
            result.get("currency", ""),
        ]
        if doc_type == "sales":
            sales_rows.append(row)
        elif doc_type == "expense":
            expense_rows.append(row)
        else:
            unrecognized.append(fpath.name)

    # --- Build the Excel report ---
    progress(1.0, desc="Generating Excel report...")
    headers = ["File", "Date", "Party", "Invoice #", "Total", "VAT", "Currency"]

    wb = openpyxl.Workbook()
    ws_sales = wb.active
    ws_sales.title = "Sales"
    ws_sales.append(headers)
    for row in sales_rows:
        ws_sales.append(row)

    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(headers)
    for row in expense_rows:
        ws_expenses.append(row)

    for ws in (ws_sales, ws_expenses):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    if unrecognized:
        ws_unrec = wb.create_sheet("Unrecognized")
        ws_unrec.append(["Filename"])
        for cell in ws_unrec[1]:
            cell.font = Font(bold=True)
        for name in unrecognized:
            ws_unrec.append([name])

    out_path = str(Path(tempfile.gettempdir()) / "accounting_report.xlsx")
    wb.save(out_path)

    total_sales = sum(r[4] for r in sales_rows if isinstance(r[4], (int, float)))
    total_expenses = sum(r[4] for r in expense_rows if isinstance(r[4], (int, float)))
    total_sales_vat = sum(r[5] for r in sales_rows if isinstance(r[5], (int, float)))
    total_expenses_vat = sum(r[5] for r in expense_rows if isinstance(r[5], (int, float)))

    summary = (
        f"Processed {len(files)} page(s)/file(s): {len(sales_rows)} sales, "
        f"{len(expense_rows)} expenses, {len(unrecognized)} unrecognized.\n"
        f"Total sales: {total_sales} (VAT: {total_sales_vat}) | "
        f"Total expenses: {total_expenses} (VAT: {total_expenses_vat})"
    )
    if unrecognized:
        summary += "\n\nNot recognized (check these manually):\n" + "\n".join(f"- {n}" for n in unrecognized)

    return out_path, summary


# --- Theme: relaxing, muted blue / white -----------------------------------
# A softer, lower-contrast palette than a typical "app blue" -- powder-blue
# page background, warm white cards, muted periwinkle accents instead of a
# saturated primary blue. Selectors below are matched against Gradio's
# actual rendered markup (verified against the installed gradio package's
# source), not guessed -- e.g. Gradio has no ".tab-nav" class; the real tab
# bar is ".tab-container" with buttons carrying role="tab" and a ".selected"
# class. Colors are written as literal hex values rather than custom CSS
# properties, since custom :root variables can fail to resolve depending on
# how/where Gradio injects this stylesheet.
#
# IMPORTANT: every *_dark variant below is pinned to match its light
# counterpart. Gradio's default Soft theme sets dark-mode text colors to
# white (e.g. block_title_text_color_dark="white") on the assumption dark
# mode also swaps in a dark background. We deliberately force a light
# background at all times (browsers/OS can request dark mode independently
# of what we want to show), so leaving the dark-mode text defaults in place
# made titles/labels render as invisible white-on-light-blue text for
# anyone with a dark-mode preference. Pinning *_dark = light value gives one
# consistent, always-legible look regardless of the visitor's OS setting.
CLARA_THEME = gr.themes.Soft(
    primary_hue=gr.themes.colors.blue,
    secondary_hue=gr.themes.colors.indigo,
    neutral_hue=gr.themes.colors.slate,
    font=[gr.themes.GoogleFont("Inter"), "ui-sans-serif", "system-ui", "sans-serif"],
).set(
    body_background_fill="#eef3fb",
    body_background_fill_dark="#eef3fb",
    background_fill_primary="#eef3fb",
    background_fill_primary_dark="#eef3fb",
    background_fill_secondary="#f6f9fd",
    background_fill_secondary_dark="#f6f9fd",
    block_background_fill="#fdfefe",
    block_background_fill_dark="#fdfefe",
    block_border_color="#dce6f5",
    block_border_color_dark="#dce6f5",
    block_border_width="1px",
    block_radius="16px",
    block_label_background_fill="#eef3fb",
    block_label_background_fill_dark="#eef3fb",
    block_label_text_color="#39527a",
    block_label_text_color_dark="#39527a",
    block_title_text_color="#2c3e63",
    block_title_text_color_dark="#2c3e63",
    body_text_color="#33415c",
    body_text_color_dark="#33415c",
    body_text_color_subdued="#6b7fa3",
    body_text_color_subdued_dark="#6b7fa3",
    button_primary_background_fill="linear-gradient(90deg, #6d8fd9, #8fb4e3)",
    button_primary_background_fill_dark="linear-gradient(90deg, #6d8fd9, #8fb4e3)",
    button_primary_background_fill_hover="linear-gradient(90deg, #5c7dc7, #7ea5d9)",
    button_primary_background_fill_hover_dark="linear-gradient(90deg, #5c7dc7, #7ea5d9)",
    button_primary_text_color="#ffffff",
    button_primary_text_color_dark="#ffffff",
    button_secondary_background_fill="#eef3fb",
    button_secondary_background_fill_dark="#eef3fb",
    button_secondary_background_fill_hover="#dce6f5",
    button_secondary_background_fill_hover_dark="#dce6f5",
    button_secondary_text_color="#39527a",
    button_secondary_text_color_dark="#39527a",
    border_color_accent="#b7cbec",
    border_color_accent_dark="#b7cbec",
    input_background_fill="#f8faff",
    input_background_fill_dark="#f8faff",
    input_border_color="#dce6f5",
    input_border_color_dark="#dce6f5",
    shadow_drop="0 2px 8px rgba(60, 90, 150, 0.05)",
    shadow_drop_lg="0 6px 20px rgba(60, 90, 150, 0.08)",
)

CLARA_CSS = """
body, .gradio-container, .dark, .dark body, .dark .gradio-container {
    background-color: #eef3fb !important;
    color: #33415c !important;
}
"""
CLARA_CSS += """
.clara-header {
    display: flex; align-items: center; gap: 16px;
    padding: 18px 24px; margin-bottom: 18px;
    background: linear-gradient(90deg, #5c7dc7 0%, #7ea5d9 55%, #a9c6e8 100%);
    border-radius: 18px;
    box-shadow: 0 6px 18px rgba(92, 125, 199, 0.18);
}
.clara-header img {
    /* Sized as an explicit 4x ratio to the title text next to it
       (h1 font-size is 1.6rem, so 4 * 1.6rem = 6.4rem here) rather than a
       fixed pixel guess, so the relationship holds if the title size ever
       changes. `!important` + `max-width: none` are required because
       Gradio's base stylesheet includes a generic
       `img { max-width: 100%; height: auto }` reset that otherwise wins
       and was letting this render at the image's native size instead of
       the intended header-icon size. */
    height: 6.4rem !important;
    width: 6.4rem !important;
    max-width: none !important;
    max-height: none !important;
    object-fit: contain;
    background: rgba(255,255,255,0.22); border-radius: 12px; padding: 6px;
    flex-shrink: 0;
}
.clara-header .clara-title h1 {
    color: #ffffff !important; margin: 0; font-size: 1.6rem; font-weight: 700;
    letter-spacing: -0.01em;
}
.clara-header .clara-title p {
    color: #eaf1fb !important; margin: 2px 0 0 0; font-size: 0.95rem;
}

/* Blocks get a soft card look against the page background */
.gradio-container .block {
    box-shadow: 0 2px 8px rgba(60, 90, 150, 0.04);
}

/* Tabs -- matches Gradio's real markup: .tabs > .tab-wrapper > .tab-container
   (role="tablist") > button[role="tab"], active state carries .selected */
.tabs > .tab-wrapper > .tab-container {
    background: #fdfefe !important;
    border-radius: 14px !important;
    padding: 4px 6px !important;
    border: 1px solid #dce6f5 !important;
    box-shadow: 0 2px 8px rgba(60, 90, 150, 0.05);
}
.tab-container button[role="tab"] {
    font-weight: 600 !important;
    color: #6b7fa3 !important;
    border-radius: 10px !important;
    border-bottom: none !important;
}
.tab-container button[role="tab"].selected {
    color: #ffffff !important;
    background: linear-gradient(90deg, #6d8fd9, #8fb4e3) !important;
}

/* Chat bubbles -- matches Gradio's real markup: .message.user / .message.bot */
.message.user {
    background: linear-gradient(90deg, #6d8fd9, #8fb4e3) !important;
    color: #ffffff !important;
    border-radius: 14px !important;
}
.message.bot {
    background: #f6f9fd !important;
    border: 1px solid #dce6f5 !important;
    border-radius: 14px !important;
}

/* Headings inside markdown blocks */
.gradio-container h1, .gradio-container h2, .gradio-container h3 {
    color: #2c3e63;
}
"""

def _logo_data_uri(path: str) -> str:
    """Base64-embeds the logo directly into the page instead of relying on
    Gradio's /file= static route (which needs the path allow-listed and can
    vary by version/config) -- this way the header logo always renders as
    long as the file exists on disk, with a graceful blank fallback if not."""
    try:
        data = base64.b64encode(Path(path).read_bytes()).decode("ascii")
        return f"data:image/png;base64,{data}"
    except Exception:
        return ""


with gr.Blocks(title="AI Server", theme=CLARA_THEME, css=CLARA_CSS) as demo:
    gr.HTML(
        f"""
        <div class="clara-header">
            <img src="{_logo_data_uri(LOGO_PATH)}" alt="logo" />
            <div class="clara-title">
                <h1> AI Server - Ibrahim Z.</h1>
                <p></p>
            </div>
        </div>
        """
    )

    with gr.Tab("Translate"):
        with gr.Row():
            with gr.Column(scale=3):
                file_in = gr.File(label="Upload document (PDF, DOCX, PPTX, image)")
                with gr.Row():
                    src = gr.Dropdown(choices=list(LANGUAGES.keys()), label="Source language", value="english")
                    tgt = gr.Dropdown(choices=list(LANGUAGES.keys()), label="Target language", value="spanish")
                    summarize = gr.Checkbox(label="Also summarize with AI")
                hebrew_doc_translate = gr.State(value=False)
                hebrew_status_translate = gr.Markdown(value="", visible=False)
                src.change(
                    set_hebrew_from_source_lang,
                    inputs=[src],
                    outputs=[hebrew_doc_translate, hebrew_status_translate],
                )
                run_btn = gr.Button("Translate")
                output_text = gr.Textbox(
                    label="Translated text", lines=20, rtl=_is_rtl(tgt.value), show_copy_button=True,
                )
                tgt.change(
                    lambda target_lang: gr.update(rtl=_is_rtl(target_lang)),
                    inputs=[tgt],
                    outputs=[output_text],
                )
                ocr_engine_out = gr.Textbox(label="OCR engine (used if the document needed OCR)", interactive=False)
                output_summary = gr.Textbox(label="Summary (if requested)", lines=5)
                run_btn.click(
                    process,
                    inputs=[file_in, src, tgt, summarize, hebrew_doc_translate],
                    outputs=[output_text, ocr_engine_out, output_summary],
                )

            with gr.Column(scale=2, visible=False) as preview_column:
                gr.Markdown("### Document preview")
                preview_image = gr.Image(
                    label="Preview", visible=False, interactive=False,
                    show_fullscreen_button=True, height=520,
                )
                preview_pdf_html = gr.HTML(visible=False)
                preview_text = gr.Textbox(
                    label="Preview", visible=False, lines=20, interactive=False,
                )
                file_in.change(
                    build_preview,
                    inputs=[file_in],
                    outputs=[preview_column, preview_image, preview_pdf_html, preview_text],
                )

    with gr.Tab("Convert to Word"):
        gr.Markdown("Upload a PDF (native or scanned) or an image — text is extracted "
                    "(with OCR if needed) and exported as a .docx file.")
        convert_file_in = gr.File(label="Upload PDF or image")
        hebrew_doc_convert = gr.Checkbox(
            label="Document is in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        convert_btn = gr.Button("Convert to Word")
        convert_output = gr.File(label="Download .docx")
        convert_ocr_engine_out = gr.Textbox(label="OCR engine actually used", interactive=False)
        convert_btn.click(
            convert_to_word,
            inputs=[convert_file_in, hebrew_doc_convert],
            outputs=[convert_output, convert_ocr_engine_out],
        )

    with gr.Tab("Chat"):
        gr.Markdown("Chat with the local AI model. Attach a PDF, DOCX, PPTX, or image "
                    "and ask questions about it — text (with OCR if needed) is extracted "
                    "and given to the model as context.\n\n"
                    "💡 Ask it to **\"make/create/generate a PowerPoint (presentation/slide "
                    "deck) about ...\"** and it will build a downloadable .pptx file — attach "
                    "a document first if you want the slides based on that document.")
        chat_hebrew_checkbox = gr.Checkbox(
            label="Attached document is in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        gr.ChatInterface(
            fn=chat_fn, type="messages", multimodal=True,
            additional_inputs=[chat_hebrew_checkbox],
            chatbot=gr.Chatbot(type="messages", height=650),
        )

    with gr.Tab("Accountant"):
        gr.Markdown(
            "Upload a **ZIP** or a **PDF** containing scans of sales invoices and expense "
            "receipts (PDF pages, PNG, JPG, TIFF, BMP). Each page/file is OCR'd and "
            "classified independently as a sale or an expense, and an Excel report is "
            "generated with both broken out on separate sheets, including totals and VAT.\n\n"
            "⚠️ **Each page must contain exactly one invoice or receipt.** Multi-page "
            "invoices (a single invoice spanning 2+ pages) are not supported — every "
            "page is treated as its own separate document, so a multi-page invoice will "
            "be split and counted incorrectly.\n\n"
            "⚠️ **The Hebrew checkbox applies to the whole batch.** If a single ZIP mixes "
            "Hebrew and non-Hebrew documents, process them in two separate batches for "
            "best accuracy."
        )
        company_name_in = gr.Textbox(
            label="Ibrahim - AI Employee",
            placeholder="e.g. Acme Corp — helps tell sales invoices from expense invoices",
        )
        zip_in = gr.File(label="Upload ZIP or PDF of invoice/receipt scans", file_types=[".zip", ".pdf"])
        hebrew_batch_in = gr.Checkbox(
            label="These documents are in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        process_btn = gr.Button("Process Invoices", variant="primary")
        report_out = gr.File(label="Download Excel Report")
        summary_out = gr.Textbox(label="Summary / Documents not recognized", lines=10)

        process_btn.click(
            process_invoices,
            inputs=[zip_in, company_name_in, hebrew_batch_in],
            outputs=[report_out, summary_out],
        )

    with gr.Tab("Attorney 32B"):
        gr.Markdown(
            "Chat with a multi-model Israeli-legal pipeline: your question is "
            "expanded into several search queries, retrieved via **hybrid "
            "BGE-M3 vector search + BM25 keyword search** (fused with "
            "Reciprocal Rank Fusion and reranked), analyzed by **Offline AI**, "
            "drafted into Hebrew by **Agent**, then "
            "**checked by Offline AI again** against the retrieved sources -- "
            "retrying retrieval up to 3 times if that check fails. See the "
            "sidebar for the sources used and the verification result. "
            "Attach a PDF, DOCX, PPTX, or image and ask questions about it — "
            "text (with OCR if needed) is extracted and given as context.\n\n"
            "🐢 Several sequential LLM calls per question (more if a "
            "verification retry fires), all on CPU -- a single answer can "
            "genuinely take a while. If that's not workable for you, try "
            "**Attorney 1.7B (Fast)** instead, which uses the simpler, "
            "single-pass retrieval pipeline.\n\n"
            "⚠️ This is not a substitute for advice from a licensed attorney. "
            "Even with verification, citations to specific laws or sections "
            "should be independently checked, and retrieval only covers what's "
            "been embedded so far (see scripts/embed_local_law_pdfs_bgem3.py)."
        )
        legal_hebrew_checkbox = gr.Checkbox(
            label="Attached document is in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        # Defined here (render=False), same reasoning as before: gr.ChatInterface
        # needs these to already exist in scope for additional_outputs, but we
        # want them placed in the right-hand sidebar, not ChatInterface's default spot.
        legal_citations_panel = gr.Markdown(_CITATIONS_PLACEHOLDER, render=False)
        legal_knesset_sources_panel = gr.Markdown(_KNESSET_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=legal_chat_fn, type="messages", multimodal=True,
                    additional_inputs=[legal_hebrew_checkbox],
                    additional_outputs=[legal_citations_panel, legal_knesset_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 📚 Citations found")
                legal_citations_panel.render()
                gr.Markdown("### 🏛️ Statutes retrieved (Knesset DB)")
                legal_knesset_sources_panel.render()

    with gr.Tab("Attorney 24B (Offline AI 14B)"):
        gr.Markdown(
            "Same pipeline as the **Attorney 32B** tab -- hybrid BGE-M3 + "
            "BM25 retrieval, Offline AI reasoning, Agent "
            "Hebrew draft, Offline AI verification with retries -- but running "
            "**a 14B reasoning model** for the reasoning/verification "
            "steps instead of the 32B model.\n\n"
            "⚡ The 14B model fits in memory alongside the 24B Agent "
            "model on a 32GB-RAM machine, so this tab doesn't pay the "
            "repeated model-reload cost the 32B tab does on every Agent<->"
            "Offline AI handoff -- meaningfully faster end-to-end. Real "
            "tradeoff: somewhat weaker legal reasoning and verification "
            "than the 32B model. Use this as your default; use **Attorney "
            "24B** (32B) when you specifically want the strongest reasoning "
            "and can afford the extra time.\n\n"
            "⚠️ This is not a substitute for advice from a licensed attorney. "
            "Even with verification, citations to specific laws or sections "
            "should be independently checked, and retrieval only covers what's "
            "been embedded so far (see scripts/embed_local_law_pdfs_bgem3.py)."
        )
        legal_14b_hebrew_checkbox = gr.Checkbox(
            label="Attached document is in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        legal_14b_citations_panel = gr.Markdown(_CITATIONS_PLACEHOLDER, render=False)
        legal_14b_sources_panel = gr.Markdown(_KNESSET_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=legal_chat_fn_deepseek14b, type="messages", multimodal=True,
                    additional_inputs=[legal_14b_hebrew_checkbox],
                    additional_outputs=[legal_14b_citations_panel, legal_14b_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 📚 Citations found")
                legal_14b_citations_panel.render()
                gr.Markdown("### 🏛️ Statutes retrieved + verification")
                legal_14b_sources_panel.render()

    with gr.Tab("Attorney 1.7B (Fast)"):
        gr.Markdown(
            "Faster variant of the Attorney tab, using a smaller **Agent** "
            "model instead of the 24B model -- much quicker replies on "
            "CPU-only hardware, since the smaller model needs a fraction of "
            "the compute and RAM the 24B does. Same Knesset-grounded RAG "
            "retrieval and safety notices as the 24B tab.\n\n"
            "Requires a one-time pull:\n"
            "`ollama pull hf.co/dicta-il/DictaLM-3.0-1.7B-Thinking-GGUF:Q4_K_M`\n\n"
            "⚡ Real tradeoff for the speed: this model is a meaningfully "
            "weaker legal reasoner than the 24B and **more prone to "
            "fabricating specific citations**. Use this for a quick rough "
            "answer or to test something; use **Attorney 32B** for anything "
            "you actually need to rely on.\n\n"
            "⚠️ This is not a substitute for advice from a licensed attorney. "
            "Citations to specific laws, sections, or cases should be independently "
            "verified — this applies even more here than on the 24B tab."
        )
        legal_fast_hebrew_checkbox = gr.Checkbox(
            label="Attached document is in Hebrew (uses Hebrew OCR instead of the default engine)"
        )
        legal_fast_citations_panel = gr.Markdown(_CITATIONS_PLACEHOLDER, render=False)
        legal_fast_knesset_sources_panel = gr.Markdown(_KNESSET_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=legal_chat_fn_fast, type="messages", multimodal=True,
                    additional_inputs=[legal_fast_hebrew_checkbox],
                    additional_outputs=[legal_fast_citations_panel, legal_fast_knesset_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 📚 Citations found")
                legal_fast_citations_panel.render()
                gr.Markdown("### 🏛️ Statutes retrieved (Knesset DB)")
                legal_fast_knesset_sources_panel.render()

    with gr.Tab("Canon AI"):
        gr.Markdown(
            "Chat about the **Codice di Diritto Canonico** (Code of Canon "
            "Law), grounded in a local vector database built from the "
            "official Italian text on vatican.va (RAG — Retrieval-"
            "Augmented Generation via ChromaDB + nomic-embed-text).\n\n"
            "⚠️ Answers are only as good as what's retrieved. This is not "
            "a substitute for a canon lawyer or the guidance of competent "
            "Church authority. Always verify against the actual canon "
            "text (linked in Sources) for anything consequential."
        )
        canon_sources_panel = gr.Markdown(_CANON_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=canon_chat_fn, type="messages",
                    additional_outputs=[canon_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 📖 Canons retrieved")
                canon_sources_panel.render()

    with gr.Tab("GDPR AI"):
        gr.Markdown(
            "Chat about the **GDPR** (Regulation (EU) 2016/679), grounded "
            "in a local vector database built from the official article "
            "text (RAG — Retrieval-Augmented Generation via ChromaDB + "
            "nomic-embed-text, answered by gpt-oss:20b).\n\n"
            "⚠️ Answers are only as good as what's retrieved. This is not "
            "a substitute for advice from a qualified data protection "
            "lawyer or DPO. Always verify against the actual article text "
            "(linked in Sources) for anything consequential."
        )
        gdpr_sources_panel = gr.Markdown(_GDPR_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=gdpr_chat_fn, type="messages",
                    additional_outputs=[gdpr_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 📜 Articles retrieved")
                gdpr_sources_panel.render()

    with gr.Tab("HIPAA AI"):
        gr.Markdown(
            "Chat about **HIPAA** (45 CFR Part 160 — General Administrative "
            "Requirements, and Part 164 — Security and Privacy, covering "
            "the Privacy, Security, and Breach Notification Rules), "
            "grounded in a local vector database built from the official "
            "regulatory text (RAG — Retrieval-Augmented Generation via "
            "ChromaDB + nomic-embed-text, answered by gpt-oss:20b).\n\n"
            "⚠️ Answers are only as good as what's retrieved. This is not "
            "a substitute for advice from a qualified healthcare "
            "compliance attorney or Privacy/Security Officer. Always "
            "verify against the actual regulatory text (linked in "
            "Sources) for anything consequential."
        )
        hipaa_sources_panel = gr.Markdown(_HIPAA_SOURCES_PLACEHOLDER, render=False)
        with gr.Row():
            with gr.Column(scale=2):
                gr.ChatInterface(
                    fn=hipaa_chat_fn, type="messages",
                    additional_outputs=[hipaa_sources_panel],
                    chatbot=gr.Chatbot(type="messages", height=650),
                )
            with gr.Column(scale=1):
                gr.Markdown("### 🏥 Sections retrieved")
                hipaa_sources_panel.render()

if __name__ == "__main__":
    demo.launch(server_name="0.0.0.0", server_port=7860)